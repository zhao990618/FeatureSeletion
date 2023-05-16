
#  ------ JADE -------#
import os

import numpy as np
import openpyxl
import pandas as pd
from random import sample
from scipy.stats import cauchy
from Filter.filter import FilterOfData
from dataProcessing.ReadDataCSV_new import ReadCSV
from dataProcessing.NonDominate import nonDominatedSort_PFAndPop, getPF
from sklearn.preprocessing import MinMaxScaler
from Classfier.KNearestNeighbors import fitnessFunction_KNN_CV
from dataProcessing.drawPlo import pltForWin


class Genetic:
    # 保存全局数据
    dataX_global = np.asarray([])
    # 保存了数据中除了class意外的其他数据
    dataX = np.asarray([])
    # 保存了类数据
    dataY = np.asarray([])
    # 特征的存放数组
    dataFeature = np.asarray([])
    # 数据的全部索引
    allDataIndex = np.asarray([])
    # 特征的数量
    dataFeatureNum = 0
    # 每一个特征的relifF评分
    scoreOfRelifF = np.asarray([])
    # 将种群中的solution提取出来
    solutionArray = []
    # 种群
    population_DE = np.asarray([])
    # 保存 在选择算子中被淘汰的个体
    inferiorPop = np.asarray([])
    # 全局档案
    globalArchive = np.asarray([])
    # 局部档案
    localArchive = np.asarray([])
    # 每一个种群的数量
    populationNum = 0
    # 种群迭代次数
    iteratorTime = 0
    # 全局保存特征数量
    globalFN = 0
    # 保存全局的数据
    globalDI = np.asarray([])
    # 精英种群通过fitness排序得到的原始前 p% 个个体得到
    p = 0
    # 实数 转 二进制
    eta = 0
    # c
    c = 0
    # 提取的最优个体数量
    bm = 0
    # 默认的自适应的F参数
    mu_F = 0
    # 默认的自适应的crossoverPro参数
    mu_CR = 0
    # 全局最优解
    globalSolution = np.asarray([])
    # 全局最优解的实数制
    globalSolutionNum = np.asarray([])
    # 全局最优解的适应度值
    globalFitness = 1
    # 全局最优的score
    globalScore = 0
    # 是否继续递归
    isRecursive = True
    # 保存数据名字
    dataName = ""

    # initialization function
    def __init__(self,dataX,dataY,dataName):
        self.dataX = dataX
        self.dataY = dataY
        self.dataName = dataName
        self.dataFeature = np.arange(self.dataX.shape[1])

        # 对数据01 归一化，将每一列的数据缩放到 [0,1]之间
        self.dataX = MinMaxScaler().fit_transform(self.dataX)
        print("进行filter，提取一部分数据")
        # filter operator
        filter = FilterOfData(dataX=self.dataX, dataY=self.dataY)

        #  compute relifF score
        self.scoreOfRelifF = filter.computerRelifFScore()
        # 进行filter过滤数据 ----  默认采用的是膝节点法
        self.scoreOfRelifF, self.dataX, self.dataFeature = filter.filter_relifF(
            scoreOfRelifF=self.scoreOfRelifF, dataX=self.dataX, dataFeature=self.dataFeature)
        # 更新 特征长度
        self.dataFeatureNum = len(self.dataFeature)
        self.allDataIndex = np.arange(self.dataX.shape[1])
        self.globalFN = self.dataFeatureNum
        self.globalDI = np.arange(self.dataX.shape[1])
        self.dataX_global = self.dataX


    # initialization parameters
    def  setParameter(self,populationNum,iteratorTime,p,eta,c,mu_F,mu_CR,bm):
        self.populationNum = populationNum
        self.iteratorTime = iteratorTime
        self.p = p
        self.eta = eta
        self.c = c
        self.mu_F = mu_F
        self.mu_CR = mu_CR
        self.bm = bm

    # 染色体
    class Chromosome:
        # 每一个个体所选择出来的特征组合   --- 二进制
        featureOfSelect = np.asarray([])
        # 每一个个体所选择出来的特征组合   --- 实数制
        featureOfSelectNum = np.asarray([])
        # 特征组合的长度
        numberOfSolution = 0
        # 特征组合的索引
        indexOfSolution = np.asarray([])
        # 所选特征长度占总长度的比例
        proportionOfFeature = 0
        # 当前的选择的fitness
        mineFitness = 0
        # 所有数据的索引
        allDataIndex = np.asarray([])
        # 选择特征索引
        selectIndex = np.asarray([])
        bestFit = 1
        bestSolution = np.asarray([])
        bestSolutionNum = np.asarray([])

        # 该个体上抛弃的最优个体
        XBest_fit = 1
        XBest_solutionNum = np.asarray([])
        # 该个体上抛弃的次优个体
        XSecondBest_fit = 1
        XSecondBest_solutionNum = np.asarray([])

        # 每一个个体变异时的步长,    # 差分进化中变异操作的F 缩放因子
        F = 0
        # 每一个个体的交叉率
        CR = 0
        # 该个体的状态标记
        flag = 0
        # 停滞状态
        SC = 0

        def __init__(self, moea, mode):
            self.allDataIndex = moea.allDataIndex
            if mode == "init":
                self.initOfChromosome(MFDE=moea)
            elif mode == "iterator":
                self.iteratorOfChromosome(MFDE=moea)

        # 自定义每一个个体的变异步长和交叉率
        def setFAndCR(self,mode):
            while  self.F <= 0 :
                # 产生该个体的F
                self.F = np.random.normal(mode.mu_F,0.1)

            if self.F > 1:
                self.F = 1

            while self.CR <= 0:
                # 产生该个体的CR
                #self.CR = np.random.standard_cauchy(mode.mu_CR,0.1)
                self.CR = cauchy.rvs(loc=mode.mu_CR, scale=0.1, size=1)[0]

            if self.CR > 1:
                self.CR = 1

        # 用于 中间产生新的个体时的初始化
        def iteratorOfChromosome(self, MFDE):
            self.featureOfSelect = np.zeros(MFDE.dataFeatureNum)
            # 自我保存，保存获取了多少的特征，为1特征的数量为多少个
            self.getNumberOfSolution()

        # 初始化
        def initOfChromosome(self, MFDE):
            self.featureOfSelect = np.zeros(MFDE.dataFeatureNum)
            self.featureOfSelectNum = np.zeros(MFDE.dataFeatureNum)
            # 随机提取 为 1 的索引
            randArray = np.asarray([])
            while len(randArray) < 2:
                randArray = sample(np.arange(MFDE.dataFeatureNum).tolist(), np.random.randint(MFDE.dataFeatureNum))
                # print(len(randArray))
            self.featureOfSelect[randArray] = 1
            for i in range(MFDE.dataFeatureNum):
                if self.featureOfSelect[i] == 1:
                    self.featureOfSelectNum[i] = np.random.uniform(MFDE.eta, 1)
                else:
                    self.featureOfSelectNum[i] = np.random.uniform(0, MFDE.eta)
            # 自我保存，保存获取了多少的特征，为1特征的数量为多少个
            self.getNumberOfSolution()
            # 计算acc
            feature_x = MFDE.dataX[:, self.featureOfSelect == 1]
            self.mineFitness = 1 - fitnessFunction_KNN_CV(findData_x=feature_x, findData_y=MFDE.dataY, CV=10)

            if self.mineFitness < self.bestFit:
                self.bestFit = self.mineFitness
                self.bestSolution = self.featureOfSelect
                self.bestSolutionNum = self.featureOfSelectNum

        # 得到选择特征数量是多少个
        def getNumberOfSolution(self):
            # 得到选择的组合里面 为1 的索引为多少个
            index = np.where(self.featureOfSelect == 1)[0]
            # 索引在dataFeature中所对应的值
            self.selectIndex = self.allDataIndex[index]
            # 存放每一个被选择的索引索引
            self.indexOfSolution = np.copy(index)
            # 得到的索引的长度是多少
            self.numberOfSolution = index.shape[0]
            # 获得比例
            self.proportionOfFeature = self.numberOfSolution / self.featureOfSelect.shape[0]

    # 初始化种群
    def initPopulation(self):
        # 种群
        for i in range(0, int(self.populationNum)):
            chromosome = self.Chromosome(moea=self, mode="init")
            self.population_DE = np.append(self.population_DE, chromosome)

    # 修改 mu_F
    def modify_F(self,set_F):
        #  Lehmer mean
        if len(set_F) == 0:
            mean_L = np.random.randn()
        else:
            mean_L = np.sum(set_F ** 2) / (np.sum(set_F) + 0.01)
        # 更新mu_F
        self.mu_F = (1 - self.c) * self.mu_F + self.c * mean_L

    # 和 mu_CR
    def modify_CR(self,set_CR):
        # arithmetic mean
        if len(set_CR) == 0:
            mean_A = np.random.randn()
        else:
            mean_A = np.mean(set_CR)
        # 更新mu_CR
        self.mu_CR = (1 - self.c) * self.mu_CR + self.c * mean_A

    # 得到当代优秀个体用于变异操作
    def getBestIndividualArray(self):
        # 用于存放每一个个体的 error
        errorArray = np.zeros(self.populationNum)
        # 进行迭代获取每一个个体的error
        for i in range(self.populationNum):
            errorArray[i] = self.population_DE[i].mineFitness

        # 进行排序
        errIndex = np.argsort(errorArray)

        # 去除最优解作为全局最优解
        if errorArray[errIndex[0]] < self.globalFitness:
            self.globalFitness = errorArray[errIndex[0]]
            self.globalSolution = self.population_DE[errIndex[0]].featureOfSelect

        # 输出最优的前三个个体
        # for i in range(3):
        #     print("acc=",1 - errorArray[errIndex[i]],"  len=",self.population_DE[errIndex[i]].numberOfSolution)


        # 最优个体数组
        bestArray = self.population_DE[errIndex[:int(self.p * self.populationNum)]]
        bestIndex = errIndex[:int(self.p * self.populationNum)]
        return bestArray,bestIndex

    # 变异操作
    def mutateOperater(self,individual_i,X_best,X_r1,X_r2):
        # 获得该个体需要进行差分的solution
        X_i = individual_i.featureOfSelectNum
        vector_1 = X_best - X_i
        vector_2 = X_r1 - X_r2
        # 进行差分
        V_i = X_i + individual_i.F * vector_1 + individual_i.F * vector_2

        # 越界修复
        V_i = np.where(V_i < 0,(0 + X_i)/2,V_i)
        V_i = np.where(V_i > 1,(1 + X_i)/2,V_i)

        return V_i

    # 交叉算子
    def crossoverOperater(self,individual_i,V_i):
        U_i = np.zeros(self.dataFeatureNum)
        # 获得了该个体的CR
        CR = individual_i.CR
        # 选择一个 位置， 保证该位置可以进行 从V_i上获得数据
        indexOfCR = np.random.randint(self.dataFeatureNum)
        # 进行交叉
        for i in range(self.dataFeatureNum):
            # 获得一个随机数
            rand_1 = np.random.randn()
            if rand_1 <= CR or i == indexOfCR:
                U_i[i] = V_i[i]
            else:
                U_i[i] = individual_i.featureOfSelectNum[i]
        return U_i

    # 选择算子---origin
    def selectOperator(self,U_i,individual_i,set_F,set_CR):
        # 将实数制转化为二进制类型
        U_i_bin = np.where(U_i > self.eta,1,0)
        # 获得 所选择的特征数据
        feature_x = self.dataX[:, U_i_bin == 1]
        tmp_u_len = len(np.where(U_i_bin == 1)[0])
        # 判断1的数量
        if tmp_u_len > 0:
            # 计算fit
            error = 1 - fitnessFunction_KNN_CV(findData_x=feature_x,
                                                findData_y=self.dataY, CV=10)
        else:
            error = 1
        # 跟新
        if error < individual_i.mineFitness:
            # 产生一个新的个体
            newIndevidual = self.Chromosome(moea=self, mode="iterator")
            # 这个新个体将遗传 原始个体的 F 和 CR
            newIndevidual.F = individual_i.F
            newIndevidual.CR = individual_i.CR
            # 将实数制类型的solution保存
            newIndevidual.featureOfSelectNum = U_i
            # 将二进制也保存起来
            newIndevidual.featureOfSelect = U_i_bin
            # 保存fitness
            newIndevidual.mineFitness = error
            # 更新
            newIndevidual.allDataIndex = individual_i.allDataIndex
            # 更新数据
            newIndevidual.getNumberOfSolution()
            # 将失败的父类保存到inferiorPop中
            self.inferiorPop = np.append(self.inferiorPop,individual_i)
            # 将成功跟新的 F 和 CR 保存
            set_F = np.append(set_F,individual_i.F)
            set_CR = np.append(set_CR,individual_i.CR)

            # 和最优解比较一下
            if newIndevidual.mineFitness < self.globalFitness:
                self.globalFitness = newIndevidual.mineFitness
                self.globalSolution = newIndevidual.featureOfSelect
                self.isRecursive = True
            elif newIndevidual.mineFitness == self.globalFitness and (
                    newIndevidual.numberOfSolution < len(np.where(self.globalSolution==1)[0])):
                self.globalSolution = newIndevidual.featureOfSelect
                self.isRecursive = True
            return newIndevidual,set_F,set_CR
        else:
            return individual_i,set_F,set_CR

    # 得到下一代
    def getNextPopulation(self):
        # 用于存放每一个成功存活的训练向量 的 F
        set_F = np.asarray([])
        # 用于存放每一个成功存活的训练向量 的 CR
        set_CR = np.asarray([])
        # 存放最优个体
        spArray = np.asarray([])
        # 获得前 p% 个的优秀个体
        bestArray,bestIndex = self.getBestIndividualArray()

        # 开始迭代
        for i in range(self.populationNum):
            # 获得第一个个体
            individual_i = self.population_DE[i]
            # 设置该个体的F和CR
            individual_i.setFAndCR(mode=self)
            # 随机选择最优个体从bestIndex选择，获得该个体在原始种群中的索引
            best_Index = i
            while best_Index == i :
                best_Index = sample(bestIndex.tolist(),1)[0]
            # 获得了第i个个体变异时的最优solution
            X_best = self.population_DE[best_Index].featureOfSelectNum
            # 获得随机个体
            randIndividualIndex = i
            # 如果选择到了 需要进行差分的个体 或者是差分个体，就需要进行重新选择
            while randIndividualIndex == i or randIndividualIndex == best_Index:
                randIndividualIndex = np.random.randint(0,self.populationNum)
            X_r1 = self.population_DE[randIndividualIndex].featureOfSelectNum
            # 将原始种群中 删除掉上述 找到的三个个体
            tempPop = np.delete(self.population_DE,[best_Index,randIndividualIndex,i])
            # tempPop种群和inferiorPop结合
            tempPop = np.concatenate([tempPop,self.inferiorPop])
            # 在这个新的种群中抽取一个个体作为最后一个变异所需的个体
            X_r2 = sample(tempPop.tolist(),1)[0].featureOfSelectNum

            # 变异
            V_i = self.mutateOperater(individual_i=individual_i,X_best=X_best
                                      ,X_r1=X_r1,X_r2=X_r2)
            # 交叉
            U_i = self.crossoverOperater(individual_i=individual_i,V_i=V_i)
            # 选择
            tmpInd, set_F, set_CR = self.selectOperator(U_i=U_i,
                        individual_i=individual_i,set_F=set_F,set_CR=set_CR)
            # 添加个体到子代种群
            spArray = np.append(spArray, tmpInd)
        # 合并
        conPop = np.concatenate((spArray, self.population_DE), axis=0)
        # 去重复
        conPop = self.deleteDuolicate(conPop=conPop)
        # 非支配排序
        if conPop.shape[0] < self.populationNum:
            # 获得需要添加的个体数量
            addNum = self.populationNum - conPop.shape[0]
            for i in range(addNum):
                # 添加新个体
                conPop = np.append(conPop, self.Chromosome(moea=self, mode="init"))
            self.population_DE = conPop
            pfArray = getPF(pop=conPop)

        else:
            pfArray, self.population_DE = nonDominatedSort_PFAndPop(
                parentPopulation=conPop, populationNum=self.populationNum)
        # 将pfArray添加到 globalArchive
        conPop = np.concatenate((pfArray, self.localArchive), axis=0)
        # 去重
        conPop = self.deleteDuolicate(conPop=conPop)
        if len(conPop) > len(self.localArchive):
            # 非支配pf
            self.localArchive = getPF(conPop)

        # 更新 mu_F mu_CR
        self.modify_F(set_F=set_F)
        self.modify_CR(set_CR=set_CR)

    # 去重复
    def deleteDuolicate(self,conPop):
        # 首先合并solution
        temp = np.asarray(conPop[0].featureOfSelect)
        for i in range(1,len(conPop)):
            temp = np.vstack((temp,conPop[i].featureOfSelect))
        # 转化为dataFrame
        df = pd.DataFrame(temp)
        # 去重
        df_dup = df.drop_duplicates()
        # 获得索引
        index_dup = df_dup.index.values
        # 提取相对应的个体
        new_pop = conPop[index_dup]
        return new_pop


    # 提取前bm个 个体，并且排序
    def getSubData(self):
        # 判断局部档案和全局档案之间的关系
        # 判断pfArray中个体的长度是否等于 self.globalArchive中个体的长度
        if self.localArchive[0].featureOfSelect.shape[0] < self.globalFN:
            for i in range(len(self.localArchive)):
                self.localArchive[i].featureOfSelect = np.zeros(self.globalFN)
                self.localArchive[i].featureOfSelect[self.localArchive[i].selectIndex] = 1
                self.localArchive[i].allDataIndex = self.globalDI
                self.localArchive[i].getNumberOfSolution()

        gl_num = self.globalArchive.shape[0]
        self.globalArchive = np.concatenate((self.globalArchive, np.asarray(self.localArchive)), axis=0)
        # 去重
        self.globalArchive = self.deleteDuolicate(conPop=self.globalArchive)
        if self.globalArchive.shape[0] > gl_num:
            self.globalArchive = getPF(pop=self.globalArchive)
        self.globalArchive = np.asarray(self.globalArchive)
        # 组合数据子集
        conArray = [[]]
        conArray[0] = np.where(self.globalArchive[0].featureOfSelect == 1)[0].tolist()
        self.bm = len(self.globalArchive)
        for i in range(1, self.bm):
            conArray.append(np.where(self.globalArchive[i].featureOfSelect == 1)[0].tolist())
        # 添加一个条件数组，当为false时则该位置所对应的的特征子集被抛弃
        isSave = np.ones(self.bm)
        # 被用于下一轮递归的数据集的特征索引
        subDataIndex = conArray[0]
        # 将该数据集中所选择的特征提取出来
        for i in range(self.bm - 1, -1, -1):
            if i == 0:
                break
            len_i = len(conArray[i])
            for j in range(i - 1, -1, -1):
                len_j = len(conArray[j])
                interArray = np.intersect1d(conArray[i], conArray[j])
                len_c_i = len(interArray)
                if (len_c_i == len_j and len_c_i <= len_i) or (len_c_i == len_i and len_c_i < len_j):
                    # 删除 第 j 个
                    isSave[i] = 0
                    break

        # 将isSave中为1的特征子集全部保存下来
        for i in range(1, self.bm):
            if isSave[i]:
                subDataIndex = np.union1d(subDataIndex, conArray[i])
        # 将这些特征都置为1
        self.dataX = self.dataX_global[:,subDataIndex]
        self.allDataIndex = self.globalDI[subDataIndex]


    # 更新全局Archive

    # 运行
    def run(self, const):
        print("第", const, "次递归")
        # 初始化函数
        self.initPopulation()
        # 进行运行
        runTime = 0
        while runTime < self.iteratorTime:
            # print("第", runTime + 1, "代")
            self.getNextPopulation()
            runTime += 1

        if self.isRecursive:
            # 提取前bm个
            self.getSubData()
            self.localArchive = np.asarray([])
            self.dataFeature = np.arange(self.dataX.shape[1])
            self.dataFeatureNum = self.dataX.shape[1]
            self.isRecursive = False
            self.population_DE = np.asarray([])
            self.inferiorPop = np.asarray([])
            self.mu_F = 0.5
            self.mu_CR = 0.5
            const += 1
            self.run(const=const)
        else:
            print("best acc = ", 1 - self.globalFitness)
            print("best len = ", len(np.where(self.globalSolution == 1)[0]))
            print()
            # 得到最终前沿
            for i in range(len(self.globalArchive)):
                print(1 - self.globalArchive[i].mineFitness, " ", self.globalArchive[i].numberOfSolution)
            print()
            print("迭代次数：", const)
            print("iteratorTime = ", self.iteratorTime)
            print("populationNum = ", self.populationNum)


if __name__ == '__main__':
    import time
    # files = os.listdir("/home/fanfan/REMODE/dataCSV")
    files = os.listdir("/home/fanfan/dataCSV/dataCSV_high")
    # files = os.listdir("/home/fanfan/dataCSV/dataCSV_temp")
    # files = os.listdir("D:/MachineLearningBackUp/dataCSV/dataCSV_high")
    # files = os.listdir("D:/MachineLearningBackUp/dataCSV/dataCSV_temp")

    for file in files:
        start = time.time()
        ducName = file.split('.')[0]
        # path_csv = "/home/fanfan/REMODE/dataCSV/" + ducName + ".csv"
        path_csv = "/home/fanfan/dataCSV/dataCSV_high/" + ducName + ".csv"
        # path_csv = "D:/MachineLearningBackUp/dataCSV/dataCSV_high/" + ducName + ".csv"
        # it_name = "RMODE_" + str(it_array[ite])
        # 写入文件的路径
        path_xlsx = "/home/fanfan/result/result_txt/compareAll/" + ducName + ".xlsx"
        # path_xlsx = "D:/MachineLearningBackUp/RecursiveCompare/MOFS_BDE/" + ducName + ".xlsx"
        # 创建xlsx
        wb = openpyxl.load_workbook(filename=path_xlsx)

        dataCsv = ReadCSV(path=path_csv)
        print("获取文件数据",file)
        dataCsv.getData()

        # 保存最终的pf
        ArchivePF = np.asarray([])

        # 循环多少次
        # iterateT = 20
        iterateT = 2

        acc = np.zeros(iterateT)
        length = np.zeros(iterateT)

        countNum = 0
        genetic = Genetic(dataX=dataCsv.dataX, dataY=dataCsv.dataY, dataName=ducName)
        for i in range(iterateT):
            genetic.setParameter(populationNum=100, iteratorTime=50, mu_F=0.5,
                                 mu_CR=0.5, eta=0.6, c=0.1, p=0.05, bm=5, rec = 1)
            genetic.run()
            print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
            print(f"all time = {time.time() - start} seconds")
            # print("ite time = ", it_array[ite])

            acc[i] = genetic.globalFitness
            length[i] = len(np.where(genetic.globalSolution == 1)[0])
            # 保存每一次循环产生的PF
            ArchivePF = np.append(ArchivePF, genetic.globalArchive)
            sheet = wb["BestAccAndLen"]
            # 添加acc到xlsx
            sheet.cell(row= 2 + i, column = 1, value=acc[i])
            # 添加len到xlsx
            sheet.cell(row= 2 + i, column = 2, value=length[i])
            wb.save(filename=path_xlsx)

            # 获得最终PF
            termPf = getPF(pop=ArchivePF)
            # 写入文件的路径
            ws_pf = wb["PF"]
            ws_pf.cell(row=1, column= countNum + 1, value="acc")
            ws_pf.cell(row=1, column= countNum + 2, value="len")
            ws_pf.cell(row=1, column= countNum + 3, value="proportion")
            ws_pf.cell(row=1, column= countNum + 4, value="recurisiveNum")
            for i in range(len(termPf)):
                # 添加acc到xlsx
                ws_pf.cell(row=2 + i, column=countNum + 1, value=termPf[i].mineFitness)
                # 添加len到xlsx
                tmp_len = len(np.where(termPf[i].featureOfSelect == 1)[0])
                ws_pf.cell(row=2 + i, column=countNum + 2, value=tmp_len)
                ws_pf.cell(row=2 + i, column=countNum + 3, value=termPf[i].proportionOfFeature)
                ws_pf.cell(row=2 + i, column=countNum + 4, value=genetic.reCount)
            wb.save(filename=path_xlsx)
            countNum += 5

            genetic.population_DE = np.asarray([])
            genetic.globalFitness = 1
            genetic.globalSolution = np.asarray([])


        #获得最终PF
        termPf = getPF(pop=ArchivePF)
        # 写入文件的路径
        global_pf = wb["gloalPF"]
        global_pf.cell(row=1, column=1, value="acc")
        global_pf.cell(row=1, column=2, value="len")
        global_pf.cell(row=1, column=3, value="proportion")
        for i in range(len(termPf)):
            # 添加acc到xlsx
            global_pf.cell(row=2 + i, column=1, value = termPf[i].mineFitness)
            # 添加len到xlsx
            tmp_len = len(np.where(termPf[i].featureOfSelect == 1)[0])
            global_pf.cell(row=2 + i, column=2, value = tmp_len)
            global_pf.cell(row=2 + i, column=3, value = termPf[i].proportionOfFeature)
            wb.save(filename=path_xlsx)

        # 向文件中写入均值
        # 写入文件的值
        # stringOfResult = str(acc.mean()) + '\t' + str(acc.std()) + '\t' + str(length.mean()) + str(length.std()) + '\n'
        # 将结果写入到文件中去
        time_std = wb["std_time"]
        time_std.cell(row=1, column=1, value="acc_std")
        time_std.cell(row=1, column=2, value="len_std")
        time_std.cell(row=1, column=3, value="time")

        time_std.cell(row=2, column=1, value=acc.std())
        time_std.cell(row=2, column=2, value=length.std())
        time_std.cell(row=2, column=3, value=(time.time() - start)/iterateT)
        wb.save(filename=path_xlsx)
        print("acc:", acc.mean(), "  std:", acc.std(), "  len:", length.mean(), "  std:", length.std())

        print(f"all time = {time.time() - start} seconds")
        print("===================================")
        print(" ")
