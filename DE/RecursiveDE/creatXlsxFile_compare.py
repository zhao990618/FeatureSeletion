import os
import numpy as np
import openpyxl

# files = os.listdir("D:/MachineLearningBackUp/dataCSV/dataCSV_high")
files = os.listdir("/home/fanfan/dataCSV/dataCSV_high")

# 创建workbook
for file in files:
    ducName = file.split('.')[0]
    # path_csv = "D:/MachineLearningBackUp/dataCSV/dataCSV_high/" + ducName + ".csv"
    path_csv = "/home/fanfan/dataCSV/dataCSV_high/" + ducName + ".csv"

    # 写入文件的路径
    # path_xlsx = "D:/MachineLearningBackUp/RecursiveDE/result/" + ducName + ".xlsx"
    # path_xlsx = "D:/MachineLearningBackUp/RecursiveCompare/MOFS_BDE/" + ducName + ".xlsx"
    # path_xlsx = "D:/MachineLearningBackUp/RecursiveCompare/MOBGA/" + ducName + ".xlsx"
    # path_xlsx = "/home/fanfan/result/result_txt/REMODE/" + ducName + ".xlsx"
    # path_xlsx = "/home/fanfan/result/result_txt/NSGA_II/" + ducName + ".xlsx"
    # path_xlsx = "/home/fanfan/result/result_txt/REDE/" + ducName + ".xlsx"
    # path_xlsx = "/home/fanfan/result/result_txt/JADE/" + ducName + ".xlsx"
    # path_xlsx = "/home/fanfan/result/result_txt/CMODE/" + ducName + ".xlsx"
    # path_xlsx = "/home/fanfan/result/result_txt/MOFS_BDE/" + ducName + ".xlsx"
    # path_xlsx = "/home/fanfan/result/result_txt/MOBGA_AOS/" + ducName + ".xlsx"
    # path_xlsx = "/home/fanfan/result/result_txt/JADE/" + ducName + ".xlsx"
    # path_xlsx = "/home/fanfan/result/result_txt/GRMOEA/" + ducName + ".xlsx"
    path_xlsx = "/home/fanfan/result/result_txt/compareAll/" + ducName + ".xlsx"
    # path_xlsx = "/home/fanfan/result/result_txt/SparseEA/" + ducName + ".xlsx"
    # 创建xlsx
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "REMODE"
    sheet.cell(row=1, column=1, value="err")
    sheet.cell(row=1, column=2, value="len")
    ws_pf2 = wb.create_sheet("CMODE")
    ws_pf3 = wb.create_sheet("GRMOEA")
    ws_pf4 = wb.create_sheet("MOBGA")
    ws_pf5 = wb.create_sheet("BDE")
    ws_pf6 = wb.create_sheet("NSGAII")
    ws_pf7 = wb.create_sheet("SparseEA")
    wb.save(filename=path_xlsx)
    print("新建Excel：" + path_xlsx + "成功")