import time

import numpy as np
from openpyxl import Workbook
import openpyxl


def creatwb(wbname):
    wb=openpyxl.Workbook()
    wb.save(filename=wbname)
    print ("新建Excel："+wbname+"成功")

# 写入excel文件中 date 数据，date是list数据类型， fields 表头
def savetoexcel(data,fields,sheetname,wbname):
    print("写入excel：")
    wb=openpyxl.load_workbook(filename=wbname)

    sheet=wb.active
    sheet.title=sheetname

    field=1
    for field in range(1,len(fields)+1):   # 写入表头
        _=sheet.cell(row=1,column=field,value=str(fields[field-1]))

    row1=1
    col1=0
    for row1 in range(2,len(data)+2):  # 写入数据
        for col1 in range(1,len(data[row1-2])+1):
            _=sheet.cell(row=row1,column=col1,value=str(data[row1-2][col1-1]))

    wb.save(filename=wbname)
    print("保存成功")


# if __name__ == '__main__':
#     wbname = "D:/MachineLearningBackUp/RecursiveDE/result/arcene.xlsx"
#     wb = openpyxl.Workbook()
#     wb.save(filename=wbname)
#     print ("新建Excel："+wbname+"成功")
#
#     # wb=openpyxl.load_workbook(filename=wbname)
#
#     sheet = wb.active
#     sheet.title="sheetname1"
#
#     # fields = np.asarray(["f1,f2"])
#     # for field in range(0,len(fields)):   # 写入表头
#     #     _=sheet.cell(row=1,column=field+1,value=str(fields[field]))
#     sheet.cell(row=1, column=1, value="acc")
#     sheet.cell(row=1, column=2, value="len")
#
#     data = np.asarray([[1,2],[3,4]])
#     row1=1
#     col1=0
#     for row1 in range(2, data.shape[0]+2):  # 写入数据
#         for col1 in range(1, data.shape[1]+1):
#             _ = sheet.cell(row=row1, column=col1, value=str(data[row1-2][col1-1]))
#     # wb.save(filename=wbname)
#     print("保存成功")
#
#     ws1 = wb.create_sheet("one")
#     ws1.cell(row=1, column=1, value="acc")
#     ws1.cell(row=1, column=2, value="len")
#
#     data = np.asarray([[1,2],[3,4]])
#     row1=1
#     col1=0
#     for row1 in range(2, data.shape[0]+2):  # 写入数据
#         for col1 in range(1, data.shape[1]+1):
#             _ = ws1.cell(row=row1, column=col1, value=str(data[row1-2][col1-1]))
#     wb.save(filename=wbname)
#     print("保存成功")


if __name__ == '__main__':
    # wbname = "D:/MachineLearningBackUp/RecursiveDE/result/arcene.xlsx"
    # wb = openpyxl.load_workbook(filename=wbname)
    # sheet = wb['BestAccAndLen']
    # sheet.cell(row=2 , column=1, value=1)
    # # 添加len到xlsx
    # sheet.cell(row=2 , column=2, value=1)
    # wb.save(wbname)
    # for i in range(1,1):
    #     print("  1")
    # print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
    a = np.asarray([0.9388,0.9388,0.9388,0.9388,0.9488,0.9388,0.95,0.9388,0.9488,0.9388])
    print(a.mean())
    print(a.std())
