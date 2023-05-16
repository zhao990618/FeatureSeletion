# competitive mechanism

import os
import random

import numpy as np
from random import sample

import openpyxl
import pandas as pd
from sklearn.preprocessing import MinMaxScaler

from Classfier.KNearestNeighbors import fitnessFunction_KNN_CV
from filer import FilterOfData
from dataProcessing.ReadDataCSV_new import ReadCSV
from dataProcessing.NonDominate import nonDominatedSort, getPF

random.seed(1)

class CMODESDE:
    # 保存了数据中除了class意外的其他数据
    dataX = np.asarray([])
    # 保存了类数据
    dataY = np.asarray([])

    # 特征的存放数组
    dataFeature = np.asarray([])
    # 特征的数量
    dataFeatureNum = 0

    # 每一个特征的relifF评分
    scoreOfRelifF = np.asarray([])

    # 将种群中的solution提取出来
    solutionArray = []
    # 任务1 的种群
    population_DE = np.asarray([])

    # 每一个种群的数量
    populationNum = 0
    # 种群迭代次数
    iteratorTime = 0
    # 精英种群通过fitness排序得到的原始前 N% 个个体得到
    eliteProb = 0

    # 交叉率
    crossoverPro = 0
    # 变异率
    mutatePro = 0
    # 差分进化中变异操作的F 缩放因子
    F = 0
    # elite number
    alpha = 0
    # limit the number of selected feature
    eta = 0

    # 全局最优解
    globalSolution = np.asarray([])
    # 全局最优解的实数制
    globalSolutionNum = np.asarray([])
    # pf
    globalArchive = np.asarray([])
    # 全局最优解的适应度值
    globalFitness = 1
    # 全局最优的score
    globalScore = 0

    # 保存数据名字
    dataName = ""

    # initialization function
    def __init__(self,dataX,dataY,ducName):
        self.dataX = dataX
        self.dataY = dataY
        # self.dataFeature = [i for i in range(self.dataX.shape[1])]
        self.dataFeature = np.arange(self.dataX.shape[1])
        self.dataFeatureNum = len(self.dataFeature)
        self.ducName=ducName
        # 对数据01 归一化，将每一列的数据缩放到 [0,1]之间
        self.dataX = MinMaxScaler().fit_transform(self.dataX)
        print("进行filter，提取一部分数据")
        # filter operator
        filter = FilterOfData(dataX=self.dataX, dataY=self.dataY)

        #  compute relifF score
        self.scoreOfRelifF = filter.computerRelifFScore()
        # 进行filter过滤数据 ----  默认采用的是膝节点法
        self.scoreOfRelifF, self.dataX, self.dataFeature = filter.filter_relifF(
            scoreOfRelifF=self.scoreOfRelifF, dataX=self.dataX, dataFeature=self.dataFeature)


    # initialization parameters
    def  setParameter(self,populationNum,iteratorTime,crossoverPro,F,alpha,eta):
        self.populationNum = populationNum
        self.iteratorTime = iteratorTime
        self.crossoverPro = crossoverPro
        self.F = F
        self.alpha = alpha
        self.eta = eta

    # 染色体
    class Chromosome:
        # 每一个个体所选择出来的特征组合   --- 二进制
        featureOfSelect = np.asarray([])
        # 每一个个体所选择出来的特征组合   --- 实数制
        featureOfSelectNum = np.asarray([])
        # 特征组合的长度
        numberOfSolution = 0
        # 特征组合的索引
        indexOfSolution = np.asarray([])
        # 所选特征长度占总长度的比例
        proportionOfFeature = 0
        # 当前的选择的fitness
        mineFitness = 0
        #
        def __init__(self, moea, mode):
            if mode == "init":
                self.initOfChromosome(MFDE=moea)
            elif mode == "iterator":
                self.iteratorOfChromosome(MFDE=moea)


        def iteratorOfChromosome(self,MFDE):
            self.featureOfSelect = np.zeros(len(MFDE.dataFeature))
            # 自我保存，保存获取了多少的特征，为1特征的数量为多少个
            self.getNumberOfSolution()
        #
        def initOfChromosome(self, MFDE):
            self.featureOfSelect = np.zeros(len(MFDE.dataFeature))
            # 首先随机初始化一个特征长度的随机数组
            rand = np.random.random(len(MFDE.dataFeature))
            for i in range(0, len(MFDE.dataFeature)):
                # 获取每一个特征上的随机数，通过这个随机数来判断初始化是否选择该特征
                if rand[i] > MFDE.eta:
                    self.featureOfSelect[i] = 1
            # 自我保存，保存获取了多少的特征，为1特征的数量为多少个
            self.getNumberOfSolution()
            self.featureOfSelectNum = rand
            # 计算acc
            feature_x = MFDE.dataX[:, self.featureOfSelect == 1]
            tmp_u_len = len(np.where(feature_x == 1)[0])
            # 判断1的数量
            if tmp_u_len > 0:
                # 计算fit
                self.mineFitness = 1 - fitnessFunction_KNN_CV(findData_x=feature_x,
                                                   findData_y=MFDE.dataY, CV=10)
            else:
                self.mineFitness = 1


        # 得到选择特征数量是多少个
        def getNumberOfSolution(self):
            # 得到选择的组合里面 为1 的索引为多少个
            index = np.where(self.featureOfSelect == 1)[0]
            # 存放每一个被选择的索引索引
            self.indexOfSolution = np.copy(index)
            # 得到的索引的长度是多少
            self.numberOfSolution = index.shape[0]
            # 获得比例
            self.proportionOfFeature = self.numberOfSolution / self.featureOfSelect.shape[0]


    # 初始化种群
    def initPopulation(self):
        # 种群
        for i in range(0, int(self.populationNum)):
            chromosome = self.Chromosome(moea=self,mode="init")
            self.population_DE = np.append(self.population_DE, chromosome)


    # obtaining elite sets by Non-Domonant sorting
    def getElietSet(self):
        fitArray = np.zeros(self.populationNum)
        lenArray = np.zeros(self.populationNum)
        for i in range(self.populationNum):
            # individual fitness
            fitArray[i] = self.population_DE[i].mineFitness
            # the length of the selected feature of the individual
            lenArray[i] = self.population_DE[i].numberOfSolution
        # determine the size of the output array based on alpha
        eliteSet,score,bestIndividual = nonDominatedSort(parentPopulation=self.population_DE,fitnessArray=fitArray,solutionlengthArray=lenArray
                         ,populationNum=self.alpha)


        # 跟新全局最优个体
        if self.globalScore <= score:
            self.globalScore = score
            self.globalFitness = bestIndividual.mineFitness
            self.globalSolution = bestIndividual.featureOfSelect

        return eliteSet

    # calculate  the fitness of each individual in population array
    def calculateFit(self,leaderPop):
        # t the individiual of the leader array compare with other individuals in the original group
        # take the minimum value of the comparison value

        # save fit of each individual
        fitValue = np.zeros(leaderPop.shape[0])
        for i in range(leaderPop.shape[0]):
            individual_i = leaderPop[i]
            compareFit = np.ones(self.populationNum)
            for j in range(self.populationNum):
                individual_j = self.population_DE[j]
                if individual_i != individual_j:
                    #  different between each objection
                    obj1_compare = individual_j.mineFitness - individual_i.mineFitness
                    if obj1_compare < 0:
                        obj1_compare = 0
                    obj2_compare = individual_j.proportionOfFeature - individual_i.proportionOfFeature
                    if obj2_compare < 0:
                        obj2_compare = 0

                    obj1_compare = obj1_compare**2
                    obj2_compare = obj2_compare**2

                    compareFit[j] = np.sqrt(obj1_compare + obj2_compare)
            # get the smallest value in the array of comparisons

            fitValue[i] = compareFit.min()
        return fitValue

    # mutate
    def mutateOperater(self,X_i,X_w,X_l):
        step_1 = self.F * (X_w - X_i)
        step_2 = self.F * (X_w - X_l)
        V_i = X_i + step_1 + step_2
        return V_i

    # crossover
    def crossoverOperator(self,X_i,V_i):
        U_i = np.ones(self.dataFeature.shape[0])
        randIndex = np.random.randint(self.dataFeature.shape[0])
        for i in range(self.dataFeature.shape[0]):
            rand = np.random.rand()
            if rand <= self.crossoverPro or i == randIndex:
                U_i[i] = V_i[i]
            else:
                U_i[i] = X_i[i]

        return U_i



    #   the main part of each generation
    def competitionBesedDEevolution(self):
        # save spring individual
        springPop = np.asarray([])

        # select eliet set
        eliteSet = self.getElietSet()

        # calculate the fit on eliteSet
        fitValue = self.calculateFit(leaderPop=eliteSet)

        # indexArray of elite
        indexArray = np.arange(self.alpha).tolist()
        for i in range(self.populationNum):
            # evolution individual
            X_i = self.population_DE[i]
            # pick two numbers at random
            randomSelectTwo = sample(indexArray,2)
            # compare the fit of the individuals corresponding to these two indices
            if fitValue[randomSelectTwo[0]] < fitValue[randomSelectTwo[1]]:
                # loser
                X_l = eliteSet[randomSelectTwo[0]]
                # winner
                X_w = eliteSet[randomSelectTwo[1]]
            else:
                # loser
                X_l = eliteSet[randomSelectTwo[1]]
                # winner
                X_w = eliteSet[randomSelectTwo[0]]

            # mutate
            V_i = self.mutateOperater(X_i=X_i.featureOfSelectNum,X_w=X_w.featureOfSelectNum,X_l=X_l.featureOfSelectNum)
            # crossover
            U_i = self.crossoverOperator(X_i=X_i.featureOfSelectNum,V_i=V_i)

            # convert real value to binary
            U_i_bin = np.where(U_i > self.eta,1,0)

            # create new individual
            newIndividual = self.Chromosome(moea=self,mode="iterator")
            # calculate acc
            feature_x = self.dataX[:, U_i_bin == 1]
            tmp_u_len = len(np.where(U_i_bin == 1)[0])
            # 判断1的数量
            if tmp_u_len > 0:
                # 计算fit
                newIndividual.mineFitness = 1 - fitnessFunction_KNN_CV(findData_x=feature_x,
                                                   findData_y=self.dataY, CV=10)
            else:
                newIndividual.mineFitness = 1
            newIndividual.featureOfSelect = U_i_bin
            newIndividual.featureOfSelectNum = U_i
            newIndividual.getNumberOfSolution()

            # append individual
            springPop = np.append(springPop,newIndividual)


        return springPop

        # 去重复

    def deleteDuolicate(self, conPop):
        # 首先合并solution
        temp = np.asarray(conPop[0].featureOfSelect)
        for i in range(1, len(conPop)):
            temp = np.vstack((temp, conPop[i].featureOfSelect))
        # 转化为dataFrame
        df = pd.DataFrame(temp)
        # 去重
        df_dup = df.drop_duplicates()
        # 获得索引
        index_dup = df_dup.index.values
        # 提取相对应的个体
        new_pop = conPop[index_dup]
        return new_pop

    #   screening individuals into the next generation
    def environmentalSelection(self,populaiton):

        fitArray = np.asarray([])
        lenArray = np.asarray([])
        popArray = np.asarray([])
        for i in range(self.populationNum):
            popArray = np.append(popArray,self.population_DE[i])
            fitArray = np.append(fitArray,self.population_DE[i].mineFitness)
            lenArray = np.append(lenArray,self.population_DE[i].numberOfSolution)

        for i in range(self.populationNum):
            popArray = np.append(popArray,populaiton[i])
            fitArray = np.append(fitArray,populaiton[i].mineFitness)
            lenArray = np.append(lenArray,populaiton[i].numberOfSolution)

        springPopulaiton,score,individual = nonDominatedSort(parentPopulation=popArray,fitnessArray=fitArray,solutionlengthArray=lenArray,
                                            populationNum=self.populationNum)
        # 跟新全局最优个体
        if self.globalScore <= score:
            self.globalScore = score
            self.globalFitness = individual.mineFitness
            self.globalSolution = individual.featureOfSelect

    def run(self):
        # 初始化种群
        self.initPopulation()
        print("进行迭代")
        # 运行次数
        runTime = 0
        while runTime < self.iteratorTime:
            # print(" 第", runTime + 1, "轮")
            springPop = self.competitionBesedDEevolution()
            self.environmentalSelection(populaiton=springPop)
            runTime += 1

        self.globalArchive = getPF(pop=self.population_DE)

if __name__ == '__main__':
    import time

    # files = os.listdir("/home/fanfan/CMODE/dataCSV")
    files = os.listdir("/home/fanfan/dataCSV/dataCSV_high")
    # files = os.listdir("D:/MachineLearningBackUp/dataCSV/dataCSV_high")
    # files = os.listdir("D:/MachineLearningBackUp/dataCSV/dataCSV_temp")

    for file in files:
        start = time.time()
        ducName = file.split('.')[0]
        path_csv = "/home/fanfan/dataCSV/dataCSV_high/" + ducName + ".csv"
        # path_csv = "D:/MachineLearningBackUp/dataCSV/dataCSV_high/" + ducName + ".csv"

        # 写入文件的路径
        path_xlsx = "/home/fanfan/result/result_txt/CMODE_SDE/" + ducName + ".xlsx"
        # path_xlsx = "D:/MachineLearningBackUp/RecursiveDE/result/" + ducName + ".xlsx"
        # 创建xlsx
        wb = openpyxl.load_workbook(filename=path_xlsx)

        dataCsv = ReadCSV(path=path_csv)
        print("获取文件数据", file)
        dataCsv.getData()

        # 保存最终的pf
        ArchivePF = np.asarray([])

        # 循环多少次
        iterateT = 10


        acc = np.zeros(iterateT)
        length = np.zeros(iterateT)

        countNum = 0
        genetic = CMODESDE(dataX=dataCsv.dataX, dataY=dataCsv.dataY, ducName=ducName)
        for i in range(iterateT):
            genetic.setParameter(populationNum=100, iteratorTime=200, crossoverPro=0.5, F=0.5, alpha=5, eta=0.6)
            genetic.run()
            genetic.population_DE = np.asarray([])
            genetic.globalScore = 0
            print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
            print(f"all time = {time.time() - start} seconds")

            acc[i] = genetic.globalFitness
            length[i] = len(np.where(genetic.globalSolution == 1)[0])
            # 保存每一次循环产生的PF
            ArchivePF = np.append(ArchivePF, genetic.globalArchive)
            sheet = wb["BestAccAndLen"]
            # 添加acc到xlsx
            sheet.cell(row=2 + i, column=1, value=acc[i])
            # 添加len到xlsx
            sheet.cell(row=2 + i, column=2, value=length[i])
            wb.save(filename=path_xlsx)

            # 获得最终PF
            termPf = getPF(pop=ArchivePF)
            # 写入文件的路径
            ws_pf = wb["PF"]
            ws_pf.cell(row=1, column=countNum + 1, value="acc")
            ws_pf.cell(row=1, column=countNum + 2, value="len")
            ws_pf.cell(row=1, column=countNum + 3, value="proportion")
            for i in range(len(termPf)):
                # 添加acc到xlsx
                ws_pf.cell(row=2 + i, column=countNum + 1, value=termPf[i].mineFitness)
                # 添加len到xlsx
                tmp_len = len(np.where(termPf[i].featureOfSelect == 1)[0])
                ws_pf.cell(row=2 + i, column=countNum + 2, value=tmp_len)
                ws_pf.cell(row=2 + i, column=countNum + 3, value=termPf[i].proportionOfFeature)
            wb.save(filename=path_xlsx)
            countNum += 4

        # 获得最终PF
        termPf = getPF(pop=ArchivePF)
        # 写入文件的路径
        global_pf = wb["gloalPF"]
        global_pf.cell(row=1, column=1, value="acc")
        global_pf.cell(row=1, column=2, value="len")
        global_pf.cell(row=1, column=3, value="proportion")
        for i in range(len(termPf)):
            # 添加acc到xlsx
            global_pf.cell(row=2 + i, column=1, value=termPf[i].mineFitness)
            # 添加len到xlsx
            tmp_len = len(np.where(termPf[i].featureOfSelect == 1)[0])
            global_pf.cell(row=2 + i, column=2, value=tmp_len)
            global_pf.cell(row=2 + i, column=3, value=termPf[i].proportionOfFeature)
            wb.save(filename=path_xlsx)

        # 向文件中写入均值
        # 写入文件的值
        stringOfResult = str(acc.mean()) + '\t' + str(acc.std()) + '\t' + str(length.mean()) + str(length.std()) + '\n'
        # 将结果写入到文件中去
        time_std = wb["std_time"]
        time_std.cell(row=1, column=1, value="acc_std")
        time_std.cell(row=1, column=2, value="len_std")
        time_std.cell(row=1, column=3, value="time")

        time_std.cell(row=2, column=1, value=acc.mean())
        time_std.cell(row=2, column=2, value=length.std())
        time_std.cell(row=2, column=3, value=(time.time() - start) / iterateT)
        wb.save(filename=path_xlsx)
        print("acc:", acc.mean(), "  std:", acc.std(), "  len:", length.mean(), "  std:", length.std())

        print(f"all time = {time.time() - start} seconds")
        print("===================================")
        print(" ")