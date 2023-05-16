# 二元差分进化 ----- 存在新的变异 ---- 纯净净化策略
import copy
import math
import os

import numpy as np
from random import sample

import openpyxl
import pandas as pd
from sklearn.preprocessing import MinMaxScaler
from Classfier.KNearestNeighbors import fitnessFunction_KNN_CV
from dataProcessing.GetParetoSolution import getParato
from dataProcessing.NonDominate import getPF, nonDominatedSort_PFAndPop
from dataProcessing.reliefF import reliefFScore
from dataProcessing.ReadDataCSV_new import ReadCSV


class MOFSBDE:
    # 保存了数据中除了class意外的其他数据
    dataX = np.asarray([])
    # 保存了类数据
    dataY = np.asarray([])
    # 数据的特征
    dataFeature = np.asarray([])
    # 特征长度
    dataFeatureLen = 0
    # 实验种群
    population_DE = np.asarray([])
    # 种群数量
    populationNum = 0
    # 迭代次数
    iteratorTime = 0
    # 交叉率
    crossoverPro = 0
    # F 进化步长
    F = 0
    # sigma
    sigma = 0
    # relifF 得分
    scoreOfRelifF = np.asarray([])
    # 文件名字
    dataName = " "
    # 进化时间
    pureTime = 0
    # 全局最优解
    globalSolution = np.asarray([])
    globalArchive = np.asarray([])
    global_acc = np.asarray([])
    global_len = np.asarray([])
    # 全局最优解的适应度值
    globalFitness = 1
    # 全局最优的score
    globalScore = 0
    # 初始化数据
    def __init__(self,dataX,dataY,dataName):
        self.dataX = dataX
        self.dataY = dataY
        self.dataName = dataName
        self.dataFeature = np.arange(self.dataX.shape[1])
        self.dataFeatureLen = len(self.dataFeature)
        # 对数据01 归一化，将每一列的数据缩放到 [0,1]之间
        self.dataX = MinMaxScaler().fit_transform(self.dataX)
        # filter
        self.filter()

    # 设置参数
    def setParameter(self, populationNum, iteratorTime, crossoverPro, F,sigma,pureTime):
        self.populationNum = populationNum
        self.iteratorTime = iteratorTime
        self.crossoverPro = crossoverPro
        self.F = F
        self.sigma = sigma
        self.pureTime = pureTime
    # filter
    def filter(self):
        # 得到relifF
        # print("得到所有特征的relifF")
        self.scoreOfRelifF = reliefFScore(self.dataX, self.dataY)
        # 拐点分割特征，选择拐点前的数据
        if len(self.dataFeature) > 1000:
            # extratRatio = findKneePoint(data=self.scoreOfRelifF)
            extratRatio = 0.03
        else:
            extratRatio = 1
        # print("通过su进行筛数据")
        self.dataX,self.scoreOfRelifF,self.dataFeature = self.extractDataOfFilter(
            extratRatio=extratRatio,dataX=self.dataX,scoreOfRelifF=self.scoreOfRelifF,dataFeature=self.dataFeature
        )
        # print()


    # 抽取数据--用于filter
    def extractDataOfFilter(self,extratRatio,dataX,scoreOfRelifF,dataFeature):
        dataLen = dataX.shape[1]
        exR = int(dataLen * extratRatio)
        # 降序排序 relifF评分
        indexOfRelifF = np.argsort(scoreOfRelifF)[::-1]
        tenpI = indexOfRelifF[:exR]
        scoreOfRelifF = scoreOfRelifF[tenpI]
        dataX = dataX[:,tenpI]
        dataFeature = dataFeature[tenpI]
        return dataX,scoreOfRelifF,dataFeature


    # 染色体
    class Chromosome:
        # 每一个个体所选择出来的特征组合   --- 二进制
        featureOfSelect = np.asarray([])
        # 特征组合的长度
        numberOfSolution = 0
        # 特征组合的索引
        indexOfSolution = np.asarray([])
        # 所选特征长度占总长度的比例
        proportionOfFeature = 0
        # 当前的选择的fitness
        mineFitness = 0
        # 交换率
        crossover = 0
        # 变异率
        mutation = 0
        # 支配等级
        p_rank = 0
        # 非支配数量
        numberOfNondominate = 0
        # 支配个体集合
        dominateSet = []
        def __init__(self, mfea):
            self.initOfChromosome(MFEA=mfea)

        def initOfChromosome(self,MFEA):
            self.featureOfSelect = np.zeros(len(MFEA.dataFeature))
            # 首先随机初始化一个特征长度的随机数组
            rand = np.random.random(len(MFEA.dataFeature))
            for i in range(0, len(MFEA.dataFeature)):
                # 获取每一个特征上的随机数，通过这个随机数来判断初始化是否选择该特征
                if rand[i] >= 0.5:
                    self.featureOfSelect[i] = 1
            # 自我保存，保存获取了多少的特征，为1特征的数量为多少个
            self.getNumberOfSolution()
            # 计算acc
            feature_x = MFEA.dataX[:, self.featureOfSelect == 1]
            acc = fitnessFunction_KNN_CV(findData_x=feature_x, findData_y=MFEA.dataY, CV=10)
            self.mineFitness = 1 - acc

        # 得到选择特征数量是多少个
        def getNumberOfSolution(self):
            # 得到选择的组合里面 为1 的索引为多少个
            index = np.where(self.featureOfSelect == 1)[0]
            # 存放每一个被选择的索引索引
            self.indexOfSolution = np.copy(index)
            # 得到的索引的长度是多少
            self.numberOfSolution = len(index)
            # 获得比例
            self.proportionOfFeature = self.numberOfSolution/len(self.featureOfSelect)

    # 初始化种群
    def initPopulation(self):
        # task1 种群
        for i in range(0, int(self.populationNum)):
            chromosome = self.Chromosome(mfea=self)
            self.population_DE = np.append(self.population_DE,chromosome)

    # DE的变异交叉选择 ---- 获得了新的下一代个体
    def mcs_DE(self):

        # 设置一个新的种群，用于存放变异交叉产生的子代
        springPopulation = np.asarray([])
        # 放置pop、acc和len
        springP = np.asarray([])
        springAcc = np.asarray([])
        springLen = np.asarray([])
        # 进行变异交叉和选择
        for a in range(self.populationNum):
            # 本轮进行变异交叉的个体
            individual_i = self.population_DE[a]
            population = np.delete(self.population_DE,a)
            # 临时数组，用于存放三个个体   ----  sample 函数 进行操作的是list
            mutationArray = np.asarray(sample(population.tolist(),3))
            # 选择三个之中最优的个体
            rankScoreArray = np.zeros(3)#len(mutationArray)
            # 保存每一个个体的error 和 len
            for i in range(3):
                rankScoreArray[i] = mutationArray[i].mineFitness * 0.9 + mutationArray[i].proportionOfFeature * 0.1
            # 对score进行升序排序
            indexOfScore = np.argsort(rankScoreArray)
            # 获得最优个体 -- 就是第一个个体
            bestIndividual = mutationArray[indexOfScore[0]]
            bestLen = bestIndividual.proportionOfFeature
            # 判断最优个体是否支配其他
            mutationArray = mutationArray[indexOfScore]
            # 支配状态
            isDominate = True
            # 判断最优个体和本轮需要进行净化个体之间的支配关系
            if (bestIndividual.mineFitness < individual_i.mineFitness and bestLen < individual_i.proportionOfFeature
            ) or (bestIndividual.mineFitness <= individual_i.mineFitness and bestLen < individual_i.proportionOfFeature
            ) or(bestIndividual.mineFitness < individual_i.mineFitness and bestLen <= individual_i.proportionOfFeature
            ):
                isDominate = True
            else:
                isDominate = False
            # 通过最优个体去进行变异,得到变异个体
            v_x = self.mutaOperater(isDominate=isDominate,mutationArray=mutationArray,individual_i = individual_i)
            # 交叉操作
            u_x = self.crossoverOperater(v_x=v_x,individual_i=individual_i)
            # 选择操作  --- 返回的数组包含了原始数组和新的数组，若是原始个体好就只包含原始个体；若是新的个体好就是包含两个
            tempP,tempAcc,tempLen = self.selectionOperation(u_x=u_x,individual_i=individual_i)
            # 添加新种群
            springP = np.append(springP,tempP)
            springAcc = np.append(springAcc,tempAcc)
            springLen = np.append(springLen,tempLen)
        # 非支配排序 --- 获得新的子代
        springPopulation = self.nonDominatedSort(
            parentPopulation=springP, fitnessArray=springAcc, solutionlengthArray=springLen)

        # print()
        return springPopulation

    # 计算每一个节点的帕累托  非支配排序   传入的是归一化后的 error 和 solutionlength
    def nonDominatedSort(self, parentPopulation, fitnessArray, solutionlengthArray):
        # 用于保存下一代解
        springsonPopulation = np.asarray([])

        # 用来保存每一个个体的函数值
        errorArray = np.asarray([])
        lenArray = np.asarray([])

        # 用于保存每一个个体的支配解集
        dominateFrontArray = []

        # 保存帕累托前沿,保存非支配数量为0的个体,
        F_i = []
        # 得到每一个个体转化后的 error  和 len
        errorArray, lenArray = self.sortFunction(error=fitnessArray, solutionLength_test=solutionlengthArray)

        ##---------------------------------- ENS  高效非支配排序 ----------------------------------------##
        # 首先对种群按照error进行一个排序，从小到大的一个排序

        # 先得到error进行排序后的索引
        sortErrorIndex = np.argsort(errorArray)
        # 将 error  population len 通过 该索引进行一个重新组合
        errorArray = errorArray[sortErrorIndex]
        parentPopulation = parentPopulation[sortErrorIndex]
        lenArray = lenArray[sortErrorIndex]
        # 对相同error的个体进行按照len的一个排序
        i = 0
        j = 0
        while i < len(errorArray) - 1:
            j = i + 1;
            if errorArray[i] == errorArray[j]:
                while j < len(errorArray) and errorArray[i] == errorArray[j]:
                    j += 1
                if (j - i > 1):
                    newArray1 = copy.deepcopy(lenArray[i:j])

                    newArray3 = copy.deepcopy(parentPopulation[i:j])
                    indexNew = np.argsort(newArray1)
                    newArray1 = newArray1[indexNew]
                    newArray3 = newArray3[indexNew]
                    for m in range(len(newArray1)):
                        lenArray[i] = newArray1[m]
                        parentPopulation[i] = newArray3[m]
                        i += 1
                    i -= 1
            i += 1
            # ==========进行迭代 --- 用的是二分查找Front============#

        # 先让 Front集合添加第一个集合
        dominateFrontArray.append([])
        # 将第一个个体添加到第一个集合
        dominateFrontArray[0].append(parentPopulation[0])

        # 第0个已经被添加到dominateFrontArray中去了，所以要从第1个开始进行比较
        for i in range(1, len(parentPopulation)):
            # 设置head 和 tail
            # head 为上界
            head = 0
            # tail 为下界
            tail = len(dominateFrontArray) - 1

            # 第 i 个 个体的 error
            error_i = errorArray[i]
            # 第 i 个 个体的 len
            len_i = lenArray[i]
            # 获得第i个个体
            individual = parentPopulation[i]

            # k 为中间的前沿
            k = math.floor((head + tail) / 2 + 0.5)

            # 进行比较
            while True:
                # 第 k 个前沿点的最后一个个体
                leastIndividual = dominateFrontArray[k][-1]
                # 找到该个体在population的中位置，才能找到其转化后的error 和len
                leastIndex = np.where(parentPopulation == leastIndividual)[0]
                # 获得最后一个个体的error
                leastError = errorArray[leastIndex]
                # 获得最后一个个体的len
                leastLen = lenArray[leastIndex]
                # 若满足下条件 则  leastIndividual  支配 第i个个体
                if (leastError < error_i and leastLen < len_i) or (
                        leastError <= error_i and leastLen < len_i) or (
                        leastError < error_i and leastLen <= len_i):
                    # 需要往下找下一个前沿
                    head = k
                    # 如过 head 和 tail 中间没有前沿了，并且tail 并不是最后一个前沿，那就把个体添加在tail所在的前沿中
                    if (tail == head + 1 and tail < len(dominateFrontArray) - 1):
                        dominateFrontArray[tail].append(individual)
                        break
                    if (tail == head + 1 and tail == len(dominateFrontArray) - 1):
                        head = tail
                        k = tail
                    # 如果 tail 为最后一个前沿，则添加一个新的前沿到dominateFrontArray
                    elif (head == len(dominateFrontArray) - 1):
                        dominateFrontArray.append([])
                        dominateFrontArray[-1].append(individual)
                        break
                    # 换到下一个二分点
                    else:
                        k = math.floor((head + tail) / 2 + 0.5)
                else:
                    # 如果没有被支配
                    # 如果只剩下三个前沿，存放在中间区域
                    if k == head + 1:
                        tail = k
                        k = head
                    elif (k == head == tail):
                        dominateFrontArray[tail].append(individual)
                        break
                    elif k == head:
                        dominateFrontArray[head].append(individual)
                        break
                    else:
                        tail = k
                        k = math.floor((head + tail) / 2 + 0.5)

        # 进行添加个体到子代
        numPop = self.populationNum
        # 计数
        count = 0
        while len(springsonPopulation) < self.populationNum:
            front_i = dominateFrontArray[count]
            resideNum = numPop - len(springsonPopulation)
            # 如果长度够，则全部添加
            if resideNum - len(front_i) >= 0:
                # 将front_i中每一个个体添加到子代种群中去
                springsonPopulation = np.append(springsonPopulation, front_i)
            else:
                # 进行拥挤度排序
                # 保存error 和 len
                crossError = np.zeros(len(front_i))
                crossLen = np.zeros(len(front_i))
                # 将前沿中每一个个体的 error和len提取出来，然后准备计算
                for k in range(0, len(front_i)):
                    crossError[k] = front_i[k].mineFitness
                    crossLen[k] = front_i[k].numberOfSolution
                crossError, crossLen = self.sortFunction(error=crossError, solutionLength_test=crossLen)
                # 计算拥挤度
                front_i, distance = self.crowdingDistance(F_i=front_i, error=crossError, lens=crossLen)
                # 从大到小排序
                sort_distance_index = np.argsort(distance)[::-1]
                # 还需要添加的个体数量
                for k in range(resideNum):
                    springsonPopulation = np.append(springsonPopulation, front_i[sort_distance_index[k]])
            count += 1

        # 保存这一代种群中的最优解
        array_score = np.zeros(len(dominateFrontArray[0]))
        for m in range(len(array_score)):
            array_score[m] = (1 - dominateFrontArray[0][m].mineFitness) * 0.9
            + (1 - dominateFrontArray[0][m].numberOfSolution / len(self.dataFeature)) * 0.1
        # 降序排序
        score_index = np.argsort(array_score)[::-1]
        # 跟新
        array_score = array_score[score_index]
        array_individual = np.asarray(dominateFrontArray[0])
        array_individual = array_individual[score_index]

        if self.globalScore <= array_score[0]:
            self.globalScore = array_score[0]
            self.globalFitness = array_individual[0].mineFitness
            self.globalSolution = array_individual[0].featureOfSelect

        self.global_acc = np.append(self.global_acc, self.globalFitness)
        self.global_len = np.append(self.global_len, self.globalSolution / len(self.dataFeature))

        return springsonPopulation

    # 拥挤度计算
    def crowdingDistance(self, F_i, error, lens):
        # 拥挤度距离
        crossedDistance = np.zeros(len(F_i))
        # 将两端的距离设置为 inf
        crossedDistance[0] = float(np.inf)
        crossedDistance[-1] = float(np.inf)

        # 计算每一个样本的拥挤度
        for i in range(1, len(F_i) - 1):
            crossedDistance[i] = crossedDistance[i] + (
                    error[i + 1] - error[i - 1]
            ) / (max(error) - min(error) + 0.001)

        for i in range(1, len(F_i) - 1):
            crossedDistance[i] = crossedDistance[i] + np.abs(
                lens[i + 1] - lens[i - 1]
            ) / (max(lens) - min(lens) + 0.01)
        return F_i, crossedDistance

    # 帕累托计算函数  ----> 1/exp(x)  1/exp(y)  [1/e,1]   -----> x == error : y == length#
    # 帕累托计算函数  ----> exp(x)  exp(y)  [1,e]   -----> x == error : y == length#
    def sortFunction(self, error, solutionLength_test):

        # 计算  e的 x次 和 e的y次
        error = np.exp(error)
        # 将长度放缩到 [0,1]之间
        solutionLength = solutionLength_test / len(self.dataFeature)
        # 计算  e的 x次 和 e的x次
        solutionLength = np.exp(solutionLength)
        return error, solutionLength
        # 变异算子

    def mutaOperater(self,isDominate,mutationArray,individual_i):
        # 特征长度
        featureLen = len(self.dataFeature)
        # 变异后的个体
        v_x = np.zeros(featureLen)
        # 最优个体
        bestIndividual = mutationArray[0]
        # 如果支配
        if isDominate:
            # 中间控制变量
            controlVariable = np.ones(featureLen) * self.sigma
        else:
            # 异或后的结果
            # 把数组转化未int类型，然后异或
            XORIndividual = mutationArray[1].featureOfSelect.astype(np.int_) ^ mutationArray[2].featureOfSelect.astype(np.int_)
            rand = np.random.uniform(0,2,featureLen)
            # 步长变量
            stepIndividual = XORIndividual * self.F * rand + self.sigma
            onesI = np.ones(featureLen)
            # 与1比较大小，选择小的，保证值不大于1
            controlVariable = np.minimum(onesI,stepIndividual)
        for i in range(featureLen):
            rand = np.random.rand()
            if controlVariable[i] < rand:
                v_x[i] = bestIndividual.featureOfSelect[i]
            else:
                v_x[i] = 1 - bestIndividual.featureOfSelect[i]
        return v_x

    # 交叉算子
    def crossoverOperater(self,v_x,individual_i):
        # 特征长度
        featureLen = len(self.dataFeature)
        # 交叉完毕后的个体
        u_x = np.zeros(featureLen)
        # 先选一个位置，保证交叉一定有一个位置是属于v_x的
        h = np.random.randint(0,featureLen)
        for i in range(featureLen):
            rand = np.random.rand()
            if i == h or rand < self.crossoverPro:
                u_x[i] = v_x[i]
            else:
                u_x[i] = individual_i.featureOfSelect[i]
        return u_x

    # 选择算子
    def selectionOperation(self,u_x,individual_i):
        # 计算acc
        feature_x = self.dataX[:, u_x == 1]
        tmp_u_len = len(np.where(u_x == 1)[0])
        # 判断1的数量
        if tmp_u_len > 0:
            # 计算fit
            acc = fitnessFunction_KNN_CV(findData_x=feature_x,findData_y=self.dataY, CV=10)
        else:
            acc = 0
        # acc = fitnessFunction_KNN_CV(findData_x=feature_x, findData_y=self.dataY, CV=10)
        u_fit = 1 - acc
        u_len = len(np.where(u_x == 1)[0])
        # 新数组
        newPop = np.asarray([])
        newPop = np.append(newPop, individual_i)
        newAcc = np.asarray([])
        newAcc = np.append(newAcc, individual_i.mineFitness)
        newLen = np.asarray([])
        newLen = np.append(newLen, individual_i.numberOfSolution)
        # 进行比较是否支配
        if ( u_fit < individual_i.mineFitness and u_len < individual_i.numberOfSolution
        ) or ( u_fit <= individual_i.mineFitness and u_len < individual_i.numberOfSolution
        ) or ( u_fit < individual_i.mineFitness and u_len <= individual_i.numberOfSolution
        ):
            # 新解支配原解
            individual_i.mineFitness = u_fit
            individual_i.featureOfSelect = u_x
            individual_i.getNumberOfSolution()

        # 原始个体支配新个体
        elif( u_fit > individual_i.mineFitness and u_len > individual_i.numberOfSolution
        ) or ( u_fit >= individual_i.mineFitness and u_len > individual_i.numberOfSolution
        ) or ( u_fit > individual_i.mineFitness and u_len >= individual_i.numberOfSolution
        ):
            pass

        # 互相不支配
        else:
            # 新的个体
            newIndividual = self.Chromosome(mfea=self)
            newIndividual.mineFitness = u_fit
            newIndividual.featureOfSelect = u_x
            newIndividual.getNumberOfSolution()
            newPop = np.append(newPop,newIndividual)
            newAcc = np.append(newAcc,u_fit)
            newLen = np.append(newLen,newIndividual.numberOfSolution)

        return newPop,newAcc,newLen

    # 特征之间的相互支配
    def featureDominate(self,p_0,p_1,individual):
        # copy一份
        x_new_0_0_f = copy.deepcopy(individual.featureOfSelect)
        x_new_0_0_f[p_1] = 0
        x_new_0_1_f = copy.deepcopy(x_new_0_0_f)
        x_new_0_1_f[p_0] = 1
        # 计算 0_0情况的acc
        feature_0_0 = self.dataX[:,x_new_0_0_f == 1]
        tmp_u_len = len(np.where(x_new_0_0_f == 1)[0])
        # 判断1的数量
        if tmp_u_len > 0:
            # 计算fit
            err_0_0 = 1 - fitnessFunction_KNN_CV(findData_x=feature_0_0,
                                                                   findData_y=self.dataY, CV=10)
        else:
            err_0_0 = 1
        # err_0_0 =1 - fitnessFunction_KNN_CV(findData_x=feature_0_0,findData_y=self.dataY,CV=10)
        # 计算 0_1情况的acc
        feature_0_1 = self.dataX[:,x_new_0_1_f == 1]
        tmp_u_len = len(np.where(x_new_0_1_f == 1)[0])
        # 判断1的数量
        if tmp_u_len > 0:
            # 计算fit
            err_0_1= 1 - fitnessFunction_KNN_CV(findData_x=feature_0_1,
                                                                   findData_y=self.dataY, CV=10)
        else:
            err_0_1 = 1
        # err_0_1 =1 - fitnessFunction_KNN_CV(findData_x=feature_0_1,findData_y=self.dataY,CV = 10)

        diff1 = np.abs(err_0_0 - individual.mineFitness)
        diff2 = np.abs(err_0_0 - err_0_1)

        if diff1 > diff2:
            return p_1
        else:
            return p_0

    # 一元纯净进化
    def oneBitPureSearch(self,population):
        # 随机选取一个个体
        randIndividual = sample(population.tolist(),1)[0]
        # 获取 1 和 0 的索引
        index_0 = np.where(randIndividual.featureOfSelect == 0)[0]
        index_1 = np.where(randIndividual.featureOfSelect == 1)[0]

        while index_0.shape[0] == 0 or index_1.shape[0] == 0:
            randIndividual = sample(population.tolist(), 1)[0]
            # 获取 1 和 0 的索引
            index_0 = np.where(randIndividual.featureOfSelect == 0)[0]
            index_1 = np.where(randIndividual.featureOfSelect == 1)[0]

        # 随机选择一个位置 --- 得到了一个为 1 和一个为 0 的索引
        position_0 = sample(index_0.tolist(),1)[0]
        position_1 = sample(index_1.tolist(),1)[0]
        # 获得支配特征 -- 从 index_0 和 index_1 中得到其中的支配特征
        dominateFeature = self.featureDominate(p_0=position_0,p_1=position_1,individual=randIndividual)
        # 指引数组 --- 按照顺序  index_0  index_1
        guideArray = np.zeros(2)
        # 判断哪一个是1
        if dominateFeature == position_0:
            guideArray[0] = 1
        else:
            guideArray[1] = 1

        # 用于保存population
        pop = np.asarray([])
        # 用于保存err
        err = np.asarray([])
        # 用于保存len
        length = np.asarray([])

        # 进行净化
        for i in range(self.populationNum):
            # 获得个体
            individual_i = population[i]
            # 复制该个体的solution
            copySolution = copy.deepcopy(individual_i.featureOfSelect)
            # 判断index_0和index_1的情况
            if copySolution[position_0] == guideArray[0] and copySolution[position_1] == guideArray[1]:
                copySolution[position_0] = guideArray[1]
                copySolution[position_1] = guideArray[0]
            else:
                copySolution[position_0] = guideArray[0]
                copySolution[position_1] = guideArray[1]

            # 新数组
            newPop = np.asarray([])
            newPop = np.append(newPop, individual_i)
            newErr = np.asarray([])
            newErr = np.append(newErr, individual_i.mineFitness)
            newLen = np.asarray([])
            newLen = np.append(newLen, individual_i.numberOfSolution)


            feature_x = self.dataX[:, copySolution == 1]
            tmp_u_len = len(np.where(copySolution == 1)[0])
            # 判断1的数量
            if tmp_u_len > 0:
                # 计算fit
                copyS_Error = 1 - fitnessFunction_KNN_CV(findData_x=feature_x, findData_y=self.dataY, CV=10)
            else:
                copyS_Error = 1

            copyS_Len = len(np.where(copySolution == 1)[0])/len(self.dataFeature)

            # 新的解支配原解
            if (copyS_Error < individual_i.mineFitness and copyS_Len < individual_i. proportionOfFeature
            ) or(copyS_Error <= individual_i.mineFitness and copyS_Len < individual_i. proportionOfFeature
            ) or(copyS_Error < individual_i.mineFitness and copyS_Len <= individual_i. proportionOfFeature
            ):
                individual_i.mineFitness = copyS_Error
                individual_i.featureOfSelect = copySolution
                individual_i.getNumberOfSolution()
            # 原解支配新解
            elif(copyS_Error > individual_i.mineFitness and copyS_Len > individual_i. proportionOfFeature
            ) or(copyS_Error >= individual_i.mineFitness and copyS_Len > individual_i. proportionOfFeature
            ) or(copyS_Error > individual_i.mineFitness and copyS_Len >= individual_i. proportionOfFeature
            ):
                pass
            # 互相不支配
            else:
                # 新的个体
                newIndividual = self.Chromosome(mfea=self)
                newIndividual.mineFitness = copyS_Error
                newIndividual.featureOfSelect = copySolution
                newIndividual.getNumberOfSolution()
                newPop = np.append(newPop, newIndividual)
                newErr = np.append(newErr, copyS_Error)
                newLen = np.append(newLen, newIndividual.numberOfSolution)
            # 将这一轮获得的支配关系个体添加进去
            pop = np.append(pop,newPop)
            err = np.append(err,newErr)
            length = np.append(length,newLen)

        # 获得非支配关系
        resultPopulation = self.nonDominatedSort(
            parentPopulation=pop, fitnessArray=err, solutionlengthArray=length)

        return resultPopulation


    # 去重复
    def deleteDuolicate(self,conPop):
        if len(conPop) <= 1:
            return  conPop
        # 首先合并solution
        temp = np.asarray(conPop[0].featureOfSelect)
        for i in range(1,len(conPop)):
            temp = np.vstack((temp,conPop[i].featureOfSelect))
        # 转化为dataFrame
        df = pd.DataFrame(temp)
        # 去重
        df_dup = df.drop_duplicates()
        # 获得索引
        index_dup = df_dup.index.values
        # 提取相对应的个体
        new_pop = conPop[index_dup]
        return new_pop


    # 运行
    def run(self):
        #初始化种群
        self.initPopulation()
        # 运行次数
        runTime = 0

        while runTime < self.iteratorTime:
            # print("第",runTime + 1,"次")
            # DE的变异交叉选择
            self.population_DE = self.mcs_DE()
            if (runTime + 1) % self.pureTime == 0:
                # 进行进化
                self.population_DE = self.oneBitPureSearch(population=self.population_DE)
            #print(f"all time = {time.time() - start} seconds")
            runTime += 1
        self.globalArchive = getPF(pop=self.population_DE)


if __name__ == '__main__':
    import time

    # files = os.listdir("/home/fanfan/MOFS_BDE/dataCSV")
    files = os.listdir("/home/fanfan/dataCSV/dataCSV_high")
    # files = os.listdir("D:/MachineLearningBackUp/dataCSV/dataCSV_high")
    # files = os.listdir("D:/MachineLearningBackUp/dataCSV/dataCSV_temp")

    for file in files:
        start = time.time()
        ducName = file.split('.')[0]
        path_csv = "/home/fanfan/dataCSV/dataCSV_high/" + ducName + ".csv"
        # path_csv = "D:/MachineLearningBackUp/dataCSV/dataCSV_high/" + ducName + ".csv"

        # 写入文件的路径
        path_xlsx = "/home/fanfan/result/result_txt/compareAll/" + ducName + ".xlsx"
        # path_xlsx = "D:/MachineLearningBackUp/RecursiveCompare/MOFS_BDE/" + ducName + ".xlsx"
        # 创建xlsx
        wb = openpyxl.load_workbook(filename=path_xlsx)

        dataCsv = ReadCSV(path=path_csv)
        print("获取文件数据", file)
        dataCsv.getData()

        # 循环多少次
        iterateT = 1
        genetic = MOFSBDE(dataX=dataCsv.dataX, dataY=dataCsv.dataY, dataName=ducName)
        for i in range(iterateT):
            # 保存最终的pf
            genetic.setParameter(populationNum=100, iteratorTime=200, crossoverPro=0.3, F=0.5, sigma=0.01,
                                 pureTime=5)
            genetic.run()
            genetic.population_DE = np.asarray([])
            genetic.globalScore = 0
            print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
            print(f"all time = {time.time() - start} seconds")

        # 写入文件的路径
        # 获得最终PF
        global_pf = wb["BDE"]
        global_pf.cell(row=1, column=1, value="acc")
        global_pf.cell(row=1, column=2, value="len")
        for i in range(len(genetic.global_acc)):
            # 添加acc到xlsx
            global_pf.cell(row=2 + i, column=1, value = genetic.global_acc[i])
            # 添加len到xlsx
            global_pf.cell(row=2 + i, column=2, value = genetic.global_len[i])
            wb.save(filename=path_xlsx)
        print(f"all time = {time.time() - start} seconds")
        print("===================================")
        print(" ")