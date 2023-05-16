import copy
import os
import time

import numpy as np
import random

import openpyxl
import pandas as pd

from Filter.filter import FilterOfData

from dataProcessing.ReadDataCSV_new import ReadCSV
from dataProcessing.NonDominate import getPF, nonDominatedSort_PFAndPop
from Classfier.KNearestNeighbors import fitnessFunction_KNN_CV
from random import sample
from sklearn.preprocessing import MinMaxScaler



# 遗传算法  ----> NSGA_II  非支配排序遗传算法--II
from dataProcessing.GetParetoSolution import getParato


class NSGAII:
    # 保存了数据中除了class意外的其他数据
    dataX = np.asarray([])
    # 保存了类数据
    dataY = np.asarray([])

    # 特征的存放数组
    dataFeature = np.asarray([])
    # 特征的数量
    dataFeatureNum = 0

    # 每一个特征的relifF评分
    scoreOfRelifF = np.asarray([])

    # 将种群中的solution提取出来
    solutionArray = []
    # 任务1 的种群
    population_EA = np.asarray([])

    # 每一个种群的数量
    populationNum = 0
    # 种群迭代次数
    iteratorTime = 0

    # 交叉率
    crossoverPro = 0
    # 变异率
    mutatePro = 0
    # 差分进化中变异操作的F 缩放因子
    F = 0
    # elite number
    alpha = 0
    # limit the number of selected feature
    eta = 0
    # 全局pf
    globalArchive = np.asarray([])
    # 全局最优解
    globalSolution = np.asarray([])
    # 全局最优解的实数制
    globalSolutionNum = np.asarray([])
    # 全局最优解的适应度值
    globalFitness = 1
    # 全局最优的score
    globalScore = 0
    global_acc = np.asarray([])
    global_len = np.asarray([])
    # 保存数据名字
    dataName = ""


    def __init__(self,dataX,dataY):
        self.dataX = dataX
        self.dataY = dataY
        self.dataFeature = np.arange(dataX.shape[1])
        self.dataFeatureNum = len(self.dataFeature)
        # 对数据01 归一化，将每一列的数据缩放到 [0,1]之间
        self.dataX = MinMaxScaler().fit_transform(self.dataX)
        print("进行filter，提取一部分数据")
        # filter operator
        filter = FilterOfData(dataX=self.dataX, dataY=self.dataY)

        #  compute relifF score
        self.scoreOfRelifF = filter.computerRelifFScore()
        # 进行filter过滤数据 ----  默认采用的是膝节点法
        self.scoreOfRelifF, self.dataX, self.dataFeature = filter.filter_relifF(
            scoreOfRelifF=self.scoreOfRelifF, dataX=self.dataX, dataFeature=self.dataFeature)

    # 设置参数
    def setParameter(self,crossoverPro,mutatePro,populationNum,iteratorTime,eta):
        self.crossoverPro = crossoverPro
        self.mutatePro = mutatePro
        self.populationNum = populationNum
        self.iteratorTime = iteratorTime
        self.eta = eta

    # 初始化染色体种群
    def initChromosomeList(self):
        for i in range(0,self.populationNum):
            chromosome = self.Chromosome(moea=self,mode="init")
            self.population_EA = np.append(self.population_EA,chromosome)

# 染色体
    class Chromosome:
        # 每一个个体所选择出来的特征组合
        featureOfSelect = np.asarray([])
        # 特征组合和长度
        numberOfSolution = 0
        # 特征组合的索引
        indexOfSolution = np.asarray([])
        # 比例
        proportionOfFeature = 0
        # 当前的选择的fitness
        mineFitness = 0
        # 交换率
        crossover = 0
        # 变异率
        mutation = 0

        # 支配等级
        p_rank = 0
        # 非支配数量
        numberOfNondominate = 0
        # 支配个体集合
        dominateSet = []



        def __init__(self,moea, mode):
            if mode == "init":
                self.initOfChromosome(MFEA=moea)
            elif mode == "iterator":
                self.iteratorOfChromosome(MFEA=moea)

        def iteratorOfChromosome(self, MFEA):
            self.featureOfSelect = np.zeros(len(MFEA.dataFeature))
            # 自我保存，保存获取了多少的特征，为1特征的数量为多少个
            self.getNumberOfSolution()
            #

        def initOfChromosome(self, MFEA):
            self.featureOfSelect = np.zeros(len(MFEA.dataFeature))
            # 首先随机初始化一个特征长度的随机数组
            rand = np.random.random(len(MFEA.dataFeature))
            for i in range(0, len(MFEA.dataFeature)):
                # 获取每一个特征上的随机数，通过这个随机数来判断初始化是否选择该特征
                if rand[i] > MFEA.eta:
                    self.featureOfSelect[i] = 1
            # 自我保存，保存获取了多少的特征，为1特征的数量为多少个
            self.getNumberOfSolution()
            # 计算acc
            feature_x = MFEA.dataX[:, self.featureOfSelect == 1]
            self.mineFitness = 1 - fitnessFunction_KNN_CV(findData_x=feature_x, findData_y=MFEA.dataY, CV=10)

        # 得到选择特征数量是多少个
        def getNumberOfSolution(self):
            # 得到选择的组合里面 为1 的索引为多少个
            self.indexOfSolution = np.where(self.featureOfSelect == 1)[0]
            # 得到的索引的长度是多少
            self.numberOfSolution = self.indexOfSolution.shape[0]
            # 获得比例
            self.proportionOfFeature = self.numberOfSolution / self.featureOfSelect.shape[0]

    #   competition
    def selectOperater(self):
        elitePopulation = np.asarray([])
        # select number
        selectNum = int(0.1 * self.populationNum)

        for i in range(self.populationNum):
            selectCompetitor = sample(self.population_EA.tolist(),selectNum)
            scoreArray = np.zeros(selectNum)
            for i in range(selectNum):
                scoreArray[i] = (selectCompetitor[i].mineFitness) * 0.9 + (selectCompetitor[i].proportionOfFeature) * 0.1
            errIndex = np.argsort(scoreArray)
            elitePopulation = np.append(elitePopulation,copy.deepcopy(selectCompetitor[errIndex[0]]))

        return elitePopulation

    # 交叉算子  使用均匀交叉，使得子代拥有极高的多样性,需要传入种群
    def crossoverOperate(self, solution1, solution2):
        # 交叉算子3  --- 均匀交叉
        # 生成一个矩阵，若是存在rand值小于该矩阵对应位置的值，那就对应的位置就要进行交叉操作
        crossProArray = np.random.random(self.dataFeature.shape[0])
        # 交叉
        for i in range(self.dataFeature.shape[0]):
            # 若是该位置两个solution所存放的值是不一样的就可以进行交叉
            if solution1[i] != solution2[i]:
                # 生成的随机数，用于判断该位置是否需要进行交叉操作
                rand = np.random.rand()
                # 若是小于交叉矩阵里的概率值
                if rand > crossProArray[i]:
                    # 进行交叉
                    temp1 = solution1[i]
                    solution2[i] = solution1[i]
                    solution1[i] = temp1

    # 变异算子
    def mutateOperator(self, individual):
        # 每一个个体都有自己的变异率
        # 变异率
        p = 1 / len(self.dataFeature)
        for index in range(0, individual.featureOfSelect.shape[0]):
            if np.random.rand() < p:
                individual.featureOfSelect[index] = self.reserveAboutMutate(individual.featureOfSelect[index])
        # 跟新个体的solution索引
        individual.getNumberOfSolution()
        # 计算两个个体 fitness
        featureX = self.dataX[:, individual.featureOfSelect == 1]
        individual.mineFitness = 1 - fitnessFunction_KNN_CV(findData_x=featureX, findData_y=self.dataY, CV=10)


    # 交换 0 和 1，用于变异
    def reserveAboutMutate(self,value):
        y = -value + 1
        return  y


    # 去重复
    def deleteDuolicate(self,conPop):
        # 首先合并solution
        temp = np.asarray(conPop[0].featureOfSelect)
        for i in range(1,len(conPop)):
            temp = np.vstack((temp,conPop[i].featureOfSelect))
        # 转化为dataFrame
        df = pd.DataFrame(temp)
        # 去重
        df_dup = df.drop_duplicates()
        # 获得索引
        index_dup = df_dup.index.values
        # 提取相对应的个体
        new_pop = conPop[index_dup]
        return new_pop


    # get spring populaiton
    def getSpringPopulation(self):

        # select
        elitePopulation = self.selectOperater()

        # cross
        crossoverOrder = np.random.permutation(self.populationNum)

        number = 0
        while number < self.populationNum:
            self.crossoverOperate(solution1=elitePopulation[crossoverOrder[number]].featureOfSelect,
                                  solution2=elitePopulation[crossoverOrder[number + 1]].featureOfSelect)
            number += 2
        # mutate
        for i in range(self.populationNum):
            self.mutateOperator(individual=elitePopulation[i])

        # combine Population
        combinePop = np.concatenate([self.population_EA,elitePopulation])

        # 非支配
        pf,self.population_EA = nonDominatedSort_PFAndPop(parentPopulation=combinePop,populationNum=self.populationNum)

        # 将pfArray添加到 globalArchive
        conPop = np.concatenate((np.asarray(pf), self.globalArchive), axis=0)
        # 去重
        if conPop.shape[0] > 1:
            conPop = self.deleteDuolicate(conPop=conPop)
        if len(conPop) > len(self.globalArchive):
            # 非支配pf
            self.globalArchive = getPF(conPop)

        # 对pf进行排序
        errArray = np.zeros(len(pf))
        for i in range(len(pf)):
            errArray[i] = pf[i].mineFitness
        # 排序
        errIndex = np.argsort(errArray)
        # print("acc = ",1 - pf[errIndex[0]].mineFitness," len = ",pf[errIndex[0]].numberOfSolution)
        # 对比
        if self.globalFitness > pf[errIndex[0]].mineFitness:
            self.globalFitness = pf[errIndex[0]].mineFitness
            self.globalSolution = pf[errIndex[0]].featureOfSelect

        self.global_acc = np.append(self.global_acc, self.globalFitness)
        self.global_len = np.append(self.global_len, self.globalSolution / len(self.dataFeature))

    # 运行
    def run(self):
        import time
        start = time.time()
        # 初始化种群
        #print("初始化种群")
        self.initChromosomeList()
        # 运行时间
        runtime = 0
        while runtime < self.iteratorTime:
            #print("第", runtime + 1, "轮")
            self.getSpringPopulation()
            runtime = runtime + 1

if __name__ == "__main__":

    files = os.listdir("/home/fanfan/dataCSV/dataCSV_high")
    # files = os.listdir("D:/MachineLearningBackUp/dataCSV/dataCSV_high")
    # files = os.listdir("D:/MachineLearningBackUp/dataCSV/dataCSV_temp")

    for file in files:
        start = time.time()
        ducName = file.split('.')[0]
        path_csv = "/home/fanfan/dataCSV/dataCSV_high/" + ducName + ".csv"
        # path_csv = "D:/MachineLearningBackUp/dataCSV/dataCSV_high/" + ducName + ".csv"

        # 写入文件的路径
        path_xlsx = "/home/fanfan/result/result_txt/compareAll/" + ducName + ".xlsx"
        # path_xlsx = "D:/MachineLearningBackUp/RecursiveDE/result/" + ducName + ".xlsx"
        # 创建xlsx
        wb = openpyxl.load_workbook(filename=path_xlsx)

        dataCsv = ReadCSV(path=path_csv)
        print("获取文件数据", file)
        print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
        dataCsv.getData()

        # 循环多少次
        iterateT = 1
        genetic = NSGAII(dataX=dataCsv.dataX, dataY=dataCsv.dataY)
        for i in range(iterateT):
            genetic.setParameter(crossoverPro=0.9,mutatePro=1/dataCsv.dataX.shape[1],populationNum=100,iteratorTime=200,eta=0.6)
            genetic.run()
            print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
            print(f"all time = {time.time() - start} seconds")

        # 获得最终PF
        global_pf = wb["NSGAII"]
        global_pf.cell(row=1, column=1, value="acc")
        global_pf.cell(row=1, column=2, value="len")
        for i in range(len(genetic.global_acc)):
            # 添加acc到xlsx
            global_pf.cell(row=2 + i, column=1, value = genetic.global_acc[i])
            # 添加len到xlsx
            global_pf.cell(row=2 + i, column=2, value = genetic.global_len[i])
            wb.save(filename=path_xlsx)

        print(f"all time = {time.time() - start} seconds")
        print("===================================")
        print(" done ")