import copy
import os
import time

import numpy as np
import random

import openpyxl
import pandas as pd

from Filter.filter import FilterOfData
from dataProcessing.ReadDataCSV_new import ReadCSV
from dataProcessing.NonDominate import nonDominatedSort_PFAndPop, getPF
from Classfier.KNearestNeighbors import fitnessFunction_KNN_CV
from random import sample
from sklearn.preprocessing import MinMaxScaler

class SparseEA:
    # 保存了数据中除了class意外的其他数据
    dataX = np.asarray([])
    # 保存了类数据
    dataY = np.asarray([])
    # 特征的存放数组
    dataFeature = np.asarray([])
    # 特征的数量
    dataFeatureNum = 0
    # 每一个特征的relifF评分
    scoreOfRelifF = np.asarray([])
    # 特征评分
    scoreOfFeature = np.asarray([])
    # 种群
    population_EA = np.asarray([])
    # 每一个种群的数量
    populationNum = 0
    # 种群迭代次数
    iteratorTime = 0
    # 交叉率
    crossoverPro = 0
    # 变异率
    mutatePro = 0
    # 全局最优解
    globalSolution = np.asarray([])
    # 全局缓存 -- 保存总迭代次数结束后的pf
    globalArchive = np.asarray([])
    # 局部缓存 -- 保存每一次迭代次数结束后的pf
    localArchive = np.asarray([])
    global_acc = np.asarray([])
    global_len = np.asarray([])
    # 全局最优解的适应度值
    globalFitness = 1
    # 保存数据名字
    dataName = ""

    def __init__(self,dataX,dataY):
        self.dataX = dataX
        self.dataY = dataY
        self.dataFeature = np.arange(dataX.shape[1])
        # 对数据01 归一化，将每一列的数据缩放到 [0,1]之间
        self.dataX = MinMaxScaler().fit_transform(self.dataX)
        print("进行filter，提取一部分数据")
        # filter operator
        filter = FilterOfData(dataX=self.dataX, dataY=self.dataY)
        #  compute relifF score
        self.scoreOfRelifF = filter.computerRelifFScore()
        # print("计算relief")
        # 进行filter过滤数据 ----
        self.scoreOfRelifF, self.dataX, self.dataFeature = filter.filter_relifF(
            scoreOfRelifF=self.scoreOfRelifF, dataX=self.dataX, dataFeature=self.dataFeature)
        # print("filter数据")
        # 获取特征数量
        self.dataFeatureNum = len(self.dataFeature)
        self.mutatePro = 1/self.dataFeatureNum
    # 设置参数
    def setParameter(self,crossoverPro,populationNum,iteratorTime):
        self.crossoverPro = crossoverPro

        self.populationNum = populationNum
        self.iteratorTime = iteratorTime

    # 计算每一个特征的分数
    def comScoreOfFeature(self):
        # 用于存放每一个特征的err
        errArray = np.zeros(self.dataFeatureNum)
        for i in range(self.dataFeatureNum):
            featureX = self.dataX[:,i]
            errArray[i] = 1 - fitnessFunction_KNN_CV(findData_x=featureX,findData_y=self.dataY,CV=10)
        # 将err进行排序
        self.scoreOfFeature = np.argsort(errArray)

    # 初始化染色体种群
    def initChromosomeList(self):
        for i in range(0,self.populationNum):
            chromosome = self.Chromosome(moea=self,mode="init")
            self.population_EA = np.append(self.population_EA,chromosome)

    class Chromosome:
        # 每一个个体所选择出来的特征组合
        featureOfSelect = np.asarray([])
        # 特征组合和长度
        numberOfSolution = 0
        # 特征组合的索引
        indexOfSolution = np.asarray([])
        # 比例
        proportionOfFeature = 0
        # 当前的选择的fitness
        mineFitness = 0
        # 支配等级
        p_rank = 0
        # 非支配数量
        numberOfNondominate = 0
        # 支配个体集合
        dominateSet = []

        def __init__(self,moea, mode):
            if mode == "init":
                self.initOfChromosome(MOEA=moea)
            elif mode == "iterator":
                self.iteratorOfChromosome(MOEA=moea)

        def iteratorOfChromosome(self, MOEA):
            self.featureOfSelect = np.zeros(len(MOEA.dataFeature))
            # 自我保存，保存获取了多少的特征，为1特征的数量为多少个
            self.getNumberOfSolution()
            #

        def initOfChromosome(self, MOEA):
            self.featureOfSelect = np.zeros(MOEA.dataFeatureNum)
            # 随机选择的特征数量
            selectNum = np.random.randint(2,MOEA.dataFeatureNum)
            # 特征序列
            featureQueue = np.arange(MOEA.dataFeatureNum)
            # 随机选择特征数量
            for i in range(selectNum):
                rF = sample(featureQueue.tolist(),2)
                if MOEA.scoreOfFeature[rF[0]] < MOEA.scoreOfFeature[rF[1]]:
                    self.featureOfSelect[rF[0]] = 1
                    fIndex = np.where(featureQueue == rF[0])[0]
                else:
                    self.featureOfSelect[rF[1]] = 1
                    fIndex = np.where(featureQueue == rF[1])[0]
                featureQueue = np.delete(featureQueue, fIndex)

            fX = MOEA.dataX[:,self.featureOfSelect == 1]
            self.mineFitness = 1 - fitnessFunction_KNN_CV(findData_x=fX,findData_y=MOEA.dataY,CV=10)
            self.getNumberOfSolution()


        # 得到选择特征数量是多少个
        def getNumberOfSolution(self):
            # 得到选择的组合里面 为1 的索引为多少个
            self.indexOfSolution = np.where(self.featureOfSelect == 1)[0]
            # 得到的索引的长度是多少
            self.numberOfSolution = self.indexOfSolution.shape[0]
            # 获得比例
            self.proportionOfFeature = self.numberOfSolution / self.featureOfSelect.shape[0]

    # 选择算子
    def selectionOperation(self):
        # 用于存放 竞标赛 留下来的个体
        sPop = np.asarray([])
        for i in range(int(self.populationNum * 2)):
            # 随机抽两个个体
            ranIndividual = sample(self.population_EA.tolist(),2)
            # 对比个体
            if ranIndividual[0].mineFitness < ranIndividual[1].mineFitness:
                sPop = np.append(sPop,ranIndividual[0])
            else:
                sPop = np.append(sPop, ranIndividual[1])

        return sPop

    # 交叉算子
    def crossoverOperator(self,S,P1,P2):
        # 三个个体 分别是 子个体 两个父个体
        rand1 = np.random.rand()
        if rand1 < 0.5:
            # 获得P1所选择特征的索引
            selectedF1 = np.where(P1.featureOfSelect == 1)[0]
            # 获得P2未选择特征的索引
            selectedF2 = np.where(P2.featureOfSelect == 0)[0]
            # 两个 索引的交集
            interSetArray = np.intersect1d(selectedF1,selectedF2)
            if interSetArray.shape[0] < 2:
                rSIndex = sample(np.union1d(selectedF1, selectedF2).tolist(), 2)
            else:
                # 随机抽取两个特征位置
                rSIndex = sample(interSetArray.tolist(),2)
            # 判断这两个特征的 score ，哪一个较差 则哪一个特征设置为 0 -- 因为 spring是复制的P1
            if self.scoreOfFeature[rSIndex[0]] < self.scoreOfFeature[rSIndex[1]]:
                S.featureOfSelect[rSIndex[1]] = 0
            else:
                S.featureOfSelect[rSIndex[0]] = 0
        else:
            # 获得P1所选择特征的索引
            selectedF1 = np.where(P1.featureOfSelect == 0)[0]
            # 获得P2未选择特征的索引
            selectedF2 = np.where(P2.featureOfSelect == 1)[0]
            # 两个 索引的交集
            interSetArray = np.intersect1d(selectedF1, selectedF2)
            if interSetArray.shape[0] < 2:
                rSIndex = sample(np.union1d(selectedF1, selectedF2).tolist(), 2)
            else:
                # 随机抽取两个特征位置
                rSIndex = sample(interSetArray.tolist(), 2)
            # 判断这两个特征的 score ，哪一个较差 则哪一个特征设置为 0 -- 因为 spring是复制的P1
            if self.scoreOfFeature[rSIndex[0]] < self.scoreOfFeature[rSIndex[1]]:
                S.featureOfSelect[rSIndex[0]] = 1
            else:
                S.featureOfSelect[rSIndex[1]] = 1
        return S
    # 变异操作
    def mutationOperator(self,S):
        rand1 = np.random.rand()
        if rand1 < 0.5:
            # 获得未选择的特征索引
            unSelected = np.where(S.featureOfSelect == 0)[0]
            if unSelected.shape[0] < 2:
                pass
            else:
                # 随机选择两个特征
                rSIndex = sample(unSelected.tolist(),2)
                # 评估特征
                if self.scoreOfFeature[rSIndex[0]] < self.scoreOfFeature[rSIndex[1]]:
                    S.featureOfSelect[rSIndex[0]] = 1
                else:
                    S.featureOfSelect[rSIndex[1]] = 1
        else:
            # 获得选择的特征索引
            unSelected = np.where(S.featureOfSelect == 1)[0]
            if unSelected.shape[0] < 2:
                pass
            else:
                # 随机选择两个特征
                rSIndex = sample(unSelected.tolist(), 2)
                # 评估特征
                if self.scoreOfFeature[rSIndex[0]] < self.scoreOfFeature[rSIndex[1]]:
                    S.featureOfSelect[rSIndex[1]] = 0
                else:
                    S.featureOfSelect[rSIndex[0]] = 0
        # 计算err
        if len(np.where(S.featureOfSelect == 1)[0]) > 0:
            fX = self.dataX[:, S.featureOfSelect == 1]
            S.mineFitness = 1 - fitnessFunction_KNN_CV(findData_x=fX, findData_y=self.dataY, CV=10)
        else:
            S.mineFitness = 1
        S.getNumberOfSolution()
        return S
    # 去重复
    def deleteDuolicate(self,conPop):
        # 首先合并solution
        temp = np.asarray(conPop[0].featureOfSelect)
        for i in range(1,len(conPop)):
            temp = np.vstack((temp,conPop[i].featureOfSelect))
        # 转化为dataFrame
        df = pd.DataFrame(temp)
        # 去重
        df_dup = df.drop_duplicates()
        # 获得索引
        index_dup = df_dup.index.values
        # 提取相对应的个体
        new_pop = conPop[index_dup]
        return new_pop

    def evolutionOfInd(self):
        # 首先选择出进化种群  -- 选择算子
        sPop = self.selectionOperation()
        # 子种群
        springPop = np.asarray([])
        # 循环
        index = 0
        while index < int(self.populationNum * 2):
            # 首先 子个体 完全复制 一个父个体的内容
            springInd = copy.deepcopy(sPop[index])
            # 其次 -- 交叉算子
            springInd = self.crossoverOperator(S=springInd,P1=sPop[index],P2=sPop[index+1])
            # 最后 -- 变异算子
            springInd = self.mutationOperator(S=springInd)
            # 添加最终个体
            springPop = np.append(springPop,springInd)
            index += 2
        # 合并子种群和父代种群
        conPop = np.concatenate((springPop,self.population_EA),axis=0)
        # 去重
        conPop = self.deleteDuolicate(conPop=conPop)
        # 非支配
        pf,self.population_EA = nonDominatedSort_PFAndPop(parentPopulation=conPop,populationNum=self.populationNum)

        # 将pfArray添加到 globalArchive
        conPop = np.concatenate((np.asarray(pf), self.globalArchive), axis=0)
        # 去重
        if conPop.shape[0] > 1:
            conPop = self.deleteDuolicate(conPop=conPop)
        if len(conPop) > len(self.globalArchive):
            # 非支配pf
            self.globalArchive = getPF(conPop)

        # 对pf进行排序
        errArray = np.zeros(len(pf))
        for i in range(len(pf)):
            errArray[i] = pf[i].mineFitness
        # 排序
        errIndex = np.argsort(errArray)
        # print("acc = ",1 - pf[errIndex[0]].mineFitness," len = ",pf[errIndex[0]].numberOfSolution)
        # 对比
        if self.globalFitness > pf[errIndex[0]].mineFitness:
            self.globalFitness = pf[errIndex[0]].mineFitness
            self.globalSolution = pf[errIndex[0]].featureOfSelect
        self.global_acc = np.append(self.global_acc, self.globalFitness)
        self.global_len = np.append(self.global_len, self.globalSolution / len(self.dataFeature))

    def run(self):
        # 计算特征score
        self.comScoreOfFeature()
        # 初始化种群
        self.initChromosomeList()
        # print("1")
        runTime = 0
        while runTime < self.iteratorTime:
            # print(" 第",runTime + 1,"轮")
            self.evolutionOfInd()
            runTime += 1

        # print("================================================")

if __name__ == "__main__":
    files = os.listdir("/home/fanfan/dataCSV/dataCSV_high")
    # files = os.listdir("D:/MachineLearningBackUp/dataCSV/dataCSV_high")
    # files = os.listdir("D:/MachineLearningBackUp/dataCSV/dataCSV_temp")

    for file in files:
        start = time.time()
        ducName = file.split('.')[0]
        path_csv = "/home/fanfan/dataCSV/dataCSV_high/" + ducName + ".csv"
        # path_csv = "D:/MachineLearningBackUp/dataCSV/dataCSV_high/" + ducName + ".csv"

        # 写入文件的路径
        path_xlsx = "/home/fanfan/result/result_txt/compareAll/" + ducName + ".xlsx"
        # path_xlsx = "D:/MachineLearningBackUp/RecursiveDE/result/" + ducName + ".xlsx"
        # path_xlsx = "D:/MachineLearningBackUp/RecursiveCompare/SparseEA/" + ducName + ".xlsx"
        # 创建xlsx
        wb = openpyxl.load_workbook(filename=path_xlsx)

        dataCsv = ReadCSV(path=path_csv)
        print("获取文件数据", file)
        dataCsv.getData()

        # 循环多少次
        iterateT = 1
        genetic = SparseEA(dataX=dataCsv.dataX, dataY=dataCsv.dataY)
        for i in range(iterateT):
            genetic.setParameter(crossoverPro=0.9,populationNum=100,iteratorTime=200)
            genetic.run()
            print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
            print(f"all time = {time.time() - start} seconds")


        global_pf = wb["SparseEA"]
        global_pf.cell(row=1, column=1, value="acc")
        global_pf.cell(row=1, column=2, value="len")
        for i in range(len(genetic.global_acc)):
            # 添加acc到xlsx
            global_pf.cell(row=2 + i, column=1, value = genetic.global_acc[i])
            # 添加len到xlsx
            global_pf.cell(row=2 + i, column=2, value = genetic.global_len[i])
            wb.save(filename=path_xlsx)


        print(f"all time = {time.time() - start} seconds")
        print("===================================")
        print(" done ")